import asyncio
import json
import os
from datetime import date
from typing import Optional

from dotenv import load_dotenv
from fastapi import BackgroundTasks, FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
import uuid
import shutil
from pathlib import Path
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

load_dotenv()

from database import get_db, init_db, row_to_dict
from ai_service import chat, chat_with_web_search, build_company_context
from push_service import init_vapid, get_public_key, send_push, send_daily_push, send_bark_push, enable_bark, disable_bark, is_bark_enabled

app = FastAPI(title="客户档案系统")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

ACCESS_PASSWORD = os.getenv("ACCESS_PASSWORD", "")


@app.middleware("http")
async def no_cache_html(request: Request, call_next):
    response = await call_next(request)
    ct = response.headers.get("content-type", "")
    path = request.url.path
    if "text/html" in ct or path == "/sw.js" or path.startswith("/app.js") or path.startswith("/style.css"):
        response.headers["Cache-Control"] = "no-store"
    return response


@app.middleware("http")
async def auth_middleware(request: Request, call_next):
    if not ACCESS_PASSWORD:
        return await call_next(request)
    if request.url.path.startswith("/api/") and request.url.path not in ("/api/auth", "/api/version") and not request.url.path.startswith("/api/calendar") and not request.url.path.startswith("/api/admin/"):
        if request.headers.get("X-Access-Token", "") != ACCESS_PASSWORD:
            return JSONResponse(status_code=401, content={"detail": "未授权"})
    return await call_next(request)


APP_VERSION = "67"

@app.get("/api/version")
async def get_version():
    return {"version": APP_VERSION}


# ── Admin ─────────────────────────────────────────────────────────────────────

@app.get("/api/admin/export-db")
async def export_db(token: str = ""):
    if ACCESS_PASSWORD and token != ACCESS_PASSWORD:
        raise HTTPException(status_code=401, detail="未授权")
    return FileResponse("client_manager.db", media_type="application/octet-stream", filename="client_manager.db")


@app.post("/api/admin/import-db")
async def import_db(request: Request, file: UploadFile = File(...)):
    if ACCESS_PASSWORD and request.headers.get("X-Access-Token", "") != ACCESS_PASSWORD:
        raise HTTPException(status_code=401, detail="未授权")
    content = await file.read()
    if len(content) < 100 or content[:6] != b"SQLite":
        raise HTTPException(status_code=400, detail="不是有效的 SQLite 文件")
    db_path = Path("client_manager.db")
    shutil.copy2(db_path, "client_manager.db.bak")
    with open(db_path, "wb") as f:
        f.write(content)
    import threading
    def _restart():
        import time
        time.sleep(1)
        os._exit(0)
    threading.Thread(target=_restart, daemon=True).start()
    return {"ok": True, "message": "数据库已替换，服务正在重启"}


@app.post("/api/admin/deploy")
async def deploy_webhook(token: str = ""):
    if ACCESS_PASSWORD and token != ACCESS_PASSWORD:
        raise HTTPException(status_code=401, detail="未授权")
    import subprocess
    cwd = str(Path(__file__).parent)
    subprocess.run(["git", "clean", "-fd", "--exclude=*.db", "--exclude=.env",
                    "--exclude=uploads", "--exclude=*.log", "--exclude=*.pem", "--exclude=*.txt"],
                   cwd=cwd, capture_output=True)
    result = subprocess.run(
        ["git", "pull", "origin", "main"],
        capture_output=True, text=True,
        cwd=cwd
    )
    output = result.stdout + result.stderr
    if result.returncode != 0:
        return {"ok": False, "output": output}
    import threading
    def _restart():
        import time
        time.sleep(1)
        os._exit(0)
    threading.Thread(target=_restart, daemon=True).start()
    return {"ok": True, "output": output}


@app.get("/admin/import")
async def admin_import_page(token: str = ""):
    if ACCESS_PASSWORD and token != ACCESS_PASSWORD:
        raise HTTPException(status_code=401, detail="未授权")
    from fastapi.responses import HTMLResponse
    html = f"""<!DOCTYPE html><html lang="zh"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>导入数据库</title>
<style>body{{font-family:sans-serif;max-width:480px;margin:60px auto;padding:0 20px}}
h2{{color:#333}}input[type=file]{{display:block;margin:16px 0}}
button{{background:#2563eb;color:#fff;border:none;padding:12px 24px;border-radius:8px;font-size:16px;cursor:pointer}}
#msg{{margin-top:16px;color:green}}.err{{color:red!important}}</style></head>
<body><h2>导入数据库</h2>
<p>选择本地下载的 <code>client_manager.db</code> 文件，上传后服务自动重启。</p>
<input type="file" id="f" accept=".db">
<button onclick="upload()">上传并替换</button>
<div id="msg"></div>
<script>
async function upload(){{
  const f=document.getElementById('f').files[0];
  if(!f){{alert('请先选择文件');return;}}
  const fd=new FormData();fd.append('file',f);
  document.getElementById('msg').textContent='上传中…';
  const r=await fetch('/api/admin/import-db',{{method:'POST',headers:{{'X-Access-Token':'{token}'}},body:fd}});
  const d=await r.json();
  const m=document.getElementById('msg');
  if(r.ok){{m.textContent=d.message+'（10秒后刷新主页面）';}}
  else{{m.textContent=d.detail;m.className='err';}}
}}
</script></body></html>"""
    return HTMLResponse(html)


# ── Auth ──────────────────────────────────────────────────────────────────────

@app.post("/api/auth")
async def auth(request: Request):
    body = await request.json()
    if body.get("password") == ACCESS_PASSWORD or not ACCESS_PASSWORD:
        return {"ok": True, "token": ACCESS_PASSWORD}
    raise HTTPException(status_code=401, detail="密码错误")


# ── Companies ─────────────────────────────────────────────────────────────────

class CompanyBody(BaseModel):
    name: str
    industry: str = ""
    level: str = "B"
    business_stage: str = "意向客户"
    last_visit_date: str = ""
    credit_limit: float = 0
    products: list = []
    tags: list = []
    notes: str = ""
    legal_rep: str = ""
    legal_rep_id: str = ""
    credit_code: str = ""
    reg_capital: str = ""
    established_date: str = ""
    reg_address: str = ""
    biz_scope: str = ""
    company_scale: str = ""
    office_address: str = ""
    employee_count: str = ""
    operating_scope: str = ""
    products_assets: list = []
    products_liabilities: list = []
    products_intermediary: list = []


@app.get("/api/companies")
def list_companies(q: str = ""):
    conn = get_db()
    if q:
        rows = conn.execute(
            """SELECT c.*,
               (SELECT name FROM contacts WHERE company_id=c.id AND is_primary=1 LIMIT 1) as primary_contact,
               (SELECT COUNT(*) FROM contacts WHERE company_id=c.id) as contact_count
               FROM companies c WHERE c.name LIKE ? ORDER BY c.level, c.updated_at DESC""",
            (f"%{q}%",),
        ).fetchall()
    else:
        rows = conn.execute(
            """SELECT c.*,
               (SELECT name FROM contacts WHERE company_id=c.id AND is_primary=1 LIMIT 1) as primary_contact,
               (SELECT COUNT(*) FROM contacts WHERE company_id=c.id) as contact_count
               FROM companies c ORDER BY c.level, c.updated_at DESC"""
        ).fetchall()
    conn.close()
    return [row_to_dict(r) for r in rows]


@app.post("/api/companies")
def create_company(body: CompanyBody):
    conn = get_db()
    cur = conn.execute(
        """INSERT INTO companies (name, industry, level, business_stage, last_visit_date,
             credit_limit, products, tags, notes,
             legal_rep, legal_rep_id, credit_code, reg_capital, established_date, reg_address, biz_scope,
             company_scale, office_address, employee_count, operating_scope,
             products_assets, products_liabilities, products_intermediary)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (body.name, body.industry, body.level, body.business_stage, body.last_visit_date,
         body.credit_limit,
         json.dumps(body.products, ensure_ascii=False),
         json.dumps(body.tags, ensure_ascii=False), body.notes,
         body.legal_rep, body.legal_rep_id, body.credit_code, body.reg_capital,
         body.established_date, body.reg_address, body.biz_scope,
         body.company_scale, body.office_address, body.employee_count, body.operating_scope,
         json.dumps(body.products_assets, ensure_ascii=False),
         json.dumps(body.products_liabilities, ensure_ascii=False),
         json.dumps(body.products_intermediary, ensure_ascii=False)),
    )
    conn.commit()
    row = conn.execute("SELECT * FROM companies WHERE id=?", (cur.lastrowid,)).fetchone()
    conn.close()
    return row_to_dict(row)


@app.get("/api/companies/{company_id}")
def get_company(company_id: int):
    conn = get_db()
    row = conn.execute("SELECT * FROM companies WHERE id=?", (company_id,)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(404, "企业不存在")
    return row_to_dict(row)


@app.put("/api/companies/{company_id}")
def update_company(company_id: int, body: CompanyBody):
    conn = get_db()
    conn.execute(
        """UPDATE companies SET name=?, industry=?, level=?, business_stage=?, last_visit_date=?,
             credit_limit=?, products=?, tags=?, notes=?,
             legal_rep=?, legal_rep_id=?, credit_code=?, reg_capital=?, established_date=?, reg_address=?, biz_scope=?,
             company_scale=?, office_address=?, employee_count=?, operating_scope=?,
             products_assets=?, products_liabilities=?, products_intermediary=?,
             updated_at=datetime('now','localtime') WHERE id=?""",
        (body.name, body.industry, body.level, body.business_stage, body.last_visit_date,
         body.credit_limit,
         json.dumps(body.products, ensure_ascii=False),
         json.dumps(body.tags, ensure_ascii=False), body.notes,
         body.legal_rep, body.legal_rep_id, body.credit_code, body.reg_capital,
         body.established_date, body.reg_address, body.biz_scope,
         body.company_scale, body.office_address, body.employee_count, body.operating_scope,
         json.dumps(body.products_assets, ensure_ascii=False),
         json.dumps(body.products_liabilities, ensure_ascii=False),
         json.dumps(body.products_intermediary, ensure_ascii=False), company_id),
    )
    conn.commit()
    row = conn.execute("SELECT * FROM companies WHERE id=?", (company_id,)).fetchone()
    conn.close()
    return row_to_dict(row)


# ── Company Products (台账) ────────────────────────────────────────────────────

class ProductBody(BaseModel):
    category: str
    product_name: str
    amount: float = 0
    credit_type: str = ""
    loan_amount: float = 0
    start_date: str = ""
    end_date: str = ""
    notes: str = ""

@app.post("/api/companies/{company_id}/equity-image")
async def upload_equity_image(company_id: int, file: UploadFile = File(...)):
    import base64, httpx as _httpx
    conn = get_db()
    row = conn.execute("SELECT name FROM companies WHERE id=?", (company_id,)).fetchone()
    if not row:
        raise HTTPException(status_code=404)
    company_name = row["name"]
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else "jpg"
    if ext not in ("jpg", "jpeg", "png", "webp", "gif"):
        ext = "jpg"
    dest = Path("uploads") / f"equity_{company_id}.{ext}"
    for old in Path("uploads").glob(f"equity_{company_id}.*"):
        old.unlink(missing_ok=True)
    content = await file.read()
    dest.write_bytes(content)

    # 视觉模型识别股权结构
    equity_data = None
    api_key = os.getenv("DEEPSEEK_API_KEY", "")
    base_url = os.getenv("DEEPSEEK_BASE_URL", "").rstrip("/")
    if api_key and "moonshot" in base_url:
        img_b64 = base64.b64encode(content).decode()
        mime = "image/jpeg" if ext in ("jpg", "jpeg") else f"image/{ext}"
        prompt = f"""这是「{company_name}」的股权结构图截图。请从图中准确提取股权关系，严格按以下JSON格式返回，不加说明和代码块：
{{
  "name": "{company_name}",
  "type": "target",
  "shareholders": [
    {{
      "name": "直接股东全称",
      "type": "company|person|state",
      "value": "持股比例如51.00%",
      "shareholders": [
        {{"name": "上层股东", "type": "person", "value": "80.00%", "shareholders": []}}
      ]
    }}
  ],
  "investments": [
    {{
      "name": "子公司全称",
      "type": "company",
      "value": "100%",
      "investments": [
        {{"name": "孙公司", "type": "company", "value": "51%", "investments": []}}
      ]
    }}
  ]
}}
只提取图中明确显示的信息，不要补充或推测。type取值：company/person/state。"""
        vision_url = (base_url if base_url.endswith("/v1") else base_url + "/v1") + "/chat/completions"
        try:
            async with _httpx.AsyncClient(timeout=45) as hc:
                r = await hc.post(vision_url,
                    headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
                    json={"model": "moonshot-v1-8k-vision-preview",
                          "messages": [{"role": "user", "content": [
                              {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{img_b64}"}},
                              {"type": "text", "text": prompt}
                          ]}],
                          "max_tokens": 3000})
            r.raise_for_status()
            raw = r.json()["choices"][0]["message"]["content"].strip()
            raw = raw.replace("```json", "").replace("```", "").strip()
            equity_data = json.loads(raw)
        except Exception:
            equity_data = None

    conn.execute("UPDATE companies SET equity_data=? WHERE id=?",
                 (json.dumps(equity_data, ensure_ascii=False) if equity_data else None, company_id))
    conn.commit()
    conn.close()
    return {"ok": True, "url": f"/uploads/equity_{company_id}.{ext}", "data": equity_data}


@app.get("/api/companies/{company_id}/equity-image-url")
def get_equity_image_url(company_id: int):
    conn = get_db()
    row = conn.execute("SELECT equity_data FROM companies WHERE id=?", (company_id,)).fetchone()
    equity_data = json.loads(row["equity_data"]) if row and row["equity_data"] else None
    for ext in ("jpg", "jpeg", "png", "webp", "gif"):
        p = Path("uploads") / f"equity_{company_id}.{ext}"
        if p.exists():
            return {"url": f"/uploads/equity_{company_id}.{ext}", "data": equity_data}
    return {"url": None, "data": equity_data}


@app.delete("/api/companies/{company_id}/equity-image")
def delete_equity_image(company_id: int):
    for ext in ("jpg", "jpeg", "png", "webp", "gif"):
        p = Path("uploads") / f"equity_{company_id}.{ext}"
        p.unlink(missing_ok=True)
    return {"ok": True}


@app.patch("/api/companies/{company_id}/equity-data")
async def update_equity_data(company_id: int, request: Request):
    body = await request.json()
    data = body.get("data")
    conn = get_db()
    conn.execute("UPDATE companies SET equity_data=? WHERE id=?",
                 (json.dumps(data, ensure_ascii=False) if data is not None else None, company_id))
    conn.commit()
    conn.close()
    return {"ok": True}


@app.get("/api/companies/{company_id}/equity-graph")
async def get_equity_graph(company_id: int, refresh: bool = False):
    conn = get_db()
    row = conn.execute("SELECT name, equity_data, equity_updated_at FROM companies WHERE id=?", (company_id,)).fetchone()
    if not row:
        raise HTTPException(status_code=404)
    name, cached, cached_at = row["name"], row["equity_data"], row["equity_updated_at"]
    if cached and not refresh:
        return json.loads(cached)
    api_key = os.getenv("DEEPSEEK_API_KEY", "")
    base_url = os.getenv("DEEPSEEK_BASE_URL", "").rstrip("/")
    if not api_key or "moonshot" not in base_url:
        raise HTTPException(status_code=400, detail="需要配置 Kimi API")
    import httpx
    api_url = (base_url if base_url.endswith("/v1") else base_url + "/v1") + "/chat/completions"
    prompt = f"""你是企业工商信息专家。请根据你掌握的工商登记数据，查询「{name}」的股权图谱。

【重要原则】
- 只填写你确实掌握、有把握的信息，不要推测或编造
- 持股比例必须是真实的登记数据，不确定则不填（value留空字符串）
- 股东姓名/公司名必须准确，不要用"某某公司"等模糊替代
- 宁可信息少但准确，不要为了完整而捏造数据

【查询内容】
1. 上穿两层：直接股东（第1层）及其主要股东或实控人（第2层）
2. 下穿两层：直接持股子公司（第1层）及其子公司（第2层）

严格按以下JSON格式返回，不加说明和代码块：
{{
  "name": "{name}",
  "type": "target",
  "shareholders": [
    {{
      "name": "直接股东全称",
      "type": "company|person|state",
      "value": "51.00%",
      "shareholders": [
        {{"name": "实控人全名", "type": "person", "value": "80.00%", "shareholders": []}}
      ]
    }}
  ],
  "investments": [
    {{
      "name": "子公司全称",
      "type": "company",
      "value": "100%",
      "investments": [
        {{"name": "孙公司全称", "type": "company", "value": "51%", "investments": []}}
      ]
    }}
  ]
}}

type取值：company=企业, person=自然人, state=国有/国资。
无数据的层级数组留空 []。若完全没有可靠信息，返回：{{"error": "未找到{name}的可靠股权信息"}}"""
    import asyncio as _aio
    payload = {"model": "moonshot-v1-8k", "messages": [{"role": "user", "content": prompt}], "max_tokens": 3000}
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    try:
        async with httpx.AsyncClient(timeout=30) as hc:
            r = await hc.post(api_url, headers=headers, json=payload)
            if r.status_code == 429:
                await _aio.sleep(12)
                r = await hc.post(api_url, headers=headers, json=payload)
        r.raise_for_status()
        raw = r.json()["choices"][0]["message"]["content"].strip()
        raw = raw.replace("```json", "").replace("```", "").strip()
        data = json.loads(raw)
    except httpx.TimeoutException:
        data = {"error": "查询超时（30s），请稍后重试"}
    except Exception as e:
        data = {"error": f"查询失败：{str(e)[:120]}"}
    conn.execute(
        "UPDATE companies SET equity_data=?, equity_updated_at=datetime('now','localtime') WHERE id=?",
        (json.dumps(data, ensure_ascii=False), company_id)
    )
    conn.commit()
    conn.close()
    return data


@app.get("/api/companies/{company_id}/products")
def list_company_products(company_id: int):
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM company_products WHERE company_id=? ORDER BY category, end_date",
        (company_id,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/api/companies/{company_id}/products")
def create_company_product(company_id: int, body: ProductBody):
    conn = get_db()
    cur = conn.execute(
        """INSERT INTO company_products (company_id, category, product_name, amount, credit_type, loan_amount, start_date, end_date, notes)
           VALUES (?,?,?,?,?,?,?,?,?)""",
        (company_id, body.category, body.product_name, body.amount,
         body.credit_type, body.loan_amount, body.start_date, body.end_date, body.notes)
    )
    conn.commit()
    row = conn.execute("SELECT * FROM company_products WHERE id=?", (cur.lastrowid,)).fetchone()
    conn.close()
    return dict(row)

@app.put("/api/companies/{company_id}/products/{product_id}")
def update_company_product(company_id: int, product_id: int, body: ProductBody):
    conn = get_db()
    conn.execute(
        """UPDATE company_products SET category=?, product_name=?, amount=?, credit_type=?, loan_amount=?, start_date=?, end_date=?, notes=?
           WHERE id=? AND company_id=?""",
        (body.category, body.product_name, body.amount, body.credit_type, body.loan_amount,
         body.start_date, body.end_date, body.notes, product_id, company_id)
    )
    conn.commit()
    row = conn.execute("SELECT * FROM company_products WHERE id=?", (product_id,)).fetchone()
    conn.close()
    return dict(row)

@app.delete("/api/companies/{company_id}/products/{product_id}")
def delete_company_product(company_id: int, product_id: int):
    conn = get_db()
    conn.execute("DELETE FROM company_products WHERE id=? AND company_id=?", (product_id, company_id))
    conn.commit()
    conn.close()
    return {"ok": True}

@app.get("/api/products/expiring")
def expiring_products(days: int = 30):
    from datetime import date, timedelta
    today = date.today().isoformat()
    deadline = (date.today() + timedelta(days=days)).isoformat()
    conn = get_db()
    rows = conn.execute(
        """SELECT cp.*, c.name as company_name FROM company_products cp
           JOIN companies c ON cp.company_id=c.id
           WHERE cp.end_date != '' AND cp.end_date BETWEEN ? AND ?
           ORDER BY cp.end_date""",
        (today, deadline)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ── Pipeline 看板 ────────────────────────────────────────────────────────────

PIPELINE_STAGES = ["意向客户", "授信申请中", "已授信", "用信中", "贷后管理", "已结清"]

@app.get("/api/pipeline")
def get_pipeline():
    conn = get_db()
    result = {}
    for stage in PIPELINE_STAGES:
        rows = conn.execute(
            """SELECT c.id, c.name, c.industry, c.level, c.business_stage, c.last_visit_date,
               (SELECT name FROM contacts WHERE company_id=c.id AND is_primary=1 LIMIT 1) as primary_contact,
               (SELECT SUM(cl.credit_amount) FROM credit_lines cl WHERE cl.company_id=c.id AND cl.status!='已结清') as total_credit
               FROM companies c WHERE c.business_stage=? ORDER BY c.level, c.updated_at DESC""",
            (stage,)
        ).fetchall()
        result[stage] = [dict(r) for r in rows]
    conn.close()
    return {"stages": PIPELINE_STAGES, "data": result}

@app.patch("/api/companies/{company_id}/stage")
def update_company_stage(company_id: int, body: dict):
    stage = body.get("stage", "")
    if stage not in PIPELINE_STAGES:
        raise HTTPException(400, "无效的业务阶段")
    conn = get_db()
    conn.execute("UPDATE companies SET business_stage=?, updated_at=datetime('now','localtime') WHERE id=?",
                 (stage, company_id))
    conn.commit()
    conn.close()
    return {"ok": True}


# ── 授信台账 ──────────────────────────────────────────────────────────────────

FACILITY_TYPES = ["最高控制额度", "组合额度", "特别授信额度", "专项额度"]

APPROVAL_TYPES = ["原始授信", "授信变更", "额度续授", "额度压缩"]

class CreditFacilityBody(BaseModel):
    facility_type: str
    parent_id: int | None = None
    approval_type: str = "原始授信"
    parent_approval_id: int | None = None
    name: str = ""
    approved_amount: float = 0
    approval_no: str = ""
    start_date: str = ""
    end_date: str = ""
    notes: str = ""

class CreditLineBody(BaseModel):
    facility_id: int | None = None
    product_name: str
    credit_type: str = ""
    credit_amount: float = 0
    used_amount: float = 0
    interest_rate: str = ""
    guarantee_type: str = ""
    start_date: str = ""
    end_date: str = ""
    status: str = "正常"
    notes: str = ""

# ── 授信设施（最高控制额度 / 组合 / 特别 / 专项）────────────────────────────

@app.get("/api/companies/{company_id}/credit-facilities")
def list_credit_facilities(company_id: int):
    conn = get_db()
    facs = [dict(r) for r in conn.execute(
        "SELECT * FROM credit_facilities WHERE company_id=? ORDER BY created_at",
        (company_id,)
    ).fetchall()]
    lines = [dict(r) for r in conn.execute(
        "SELECT * FROM credit_lines WHERE company_id=? ORDER BY end_date, created_at DESC",
        (company_id,)
    ).fetchall()]
    conn.close()
    return {"facilities": facs, "lines": lines}

@app.post("/api/companies/{company_id}/credit-facilities")
def create_credit_facility(company_id: int, body: CreditFacilityBody):
    if body.facility_type not in FACILITY_TYPES:
        raise HTTPException(400, "无效的额度类型")
    conn = get_db()
    # 若是新变更批复（最高控制额度级别），将原批复标为已变更
    if body.facility_type == "最高控制额度" and body.parent_approval_id:
        conn.execute(
            "UPDATE credit_facilities SET is_active=0 WHERE id=?",
            (body.parent_approval_id,)
        )
    cur = conn.execute(
        """INSERT INTO credit_facilities (company_id, facility_type, parent_id, approval_type,
             parent_approval_id, name, approved_amount, approval_no, start_date, end_date, notes, is_active)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,1)""",
        (company_id, body.facility_type, body.parent_id, body.approval_type,
         body.parent_approval_id, body.name,
         body.approved_amount, body.approval_no, body.start_date, body.end_date, body.notes)
    )
    conn.commit()
    row = conn.execute("SELECT * FROM credit_facilities WHERE id=?", (cur.lastrowid,)).fetchone()
    conn.close()
    return dict(row)

@app.put("/api/credit-facilities/{fac_id}")
def update_credit_facility(fac_id: int, body: CreditFacilityBody):
    conn = get_db()
    conn.execute(
        """UPDATE credit_facilities SET facility_type=?, parent_id=?, approval_type=?,
             parent_approval_id=?, name=?, approved_amount=?, approval_no=?,
             start_date=?, end_date=?, notes=?
           WHERE id=?""",
        (body.facility_type, body.parent_id, body.approval_type,
         body.parent_approval_id, body.name,
         body.approved_amount, body.approval_no, body.start_date, body.end_date, body.notes,
         fac_id)
    )
    conn.commit()
    row = conn.execute("SELECT * FROM credit_facilities WHERE id=?", (fac_id,)).fetchone()
    conn.close()
    return dict(row)

@app.patch("/api/credit-facilities/{fac_id}/activate")
def activate_credit_facility(fac_id: int):
    """将某批复设为有效，同批次其他最高控制额度设为已变更。"""
    conn = get_db()
    row = conn.execute("SELECT * FROM credit_facilities WHERE id=?", (fac_id,)).fetchone()
    if not row:
        raise HTTPException(404)
    if row["facility_type"] == "最高控制额度":
        conn.execute(
            """UPDATE credit_facilities SET is_active=0
               WHERE company_id=? AND facility_type='最高控制额度' AND id!=?""",
            (row["company_id"], fac_id)
        )
    conn.execute("UPDATE credit_facilities SET is_active=1 WHERE id=?", (fac_id,))
    conn.commit()
    conn.close()
    return {"ok": True}

@app.delete("/api/credit-facilities/{fac_id}")
def delete_credit_facility(fac_id: int):
    conn = get_db()
    conn.execute("DELETE FROM credit_facilities WHERE id=?", (fac_id,))
    conn.commit()
    conn.close()
    return {"ok": True}

# ── 具体业务 ──────────────────────────────────────────────────────────────────

@app.get("/api/companies/{company_id}/credit-lines")
def list_credit_lines(company_id: int):
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM credit_lines WHERE company_id=? ORDER BY facility_id, end_date, created_at DESC",
        (company_id,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/api/companies/{company_id}/credit-lines")
def create_credit_line(company_id: int, body: CreditLineBody):
    conn = get_db()
    cur = conn.execute(
        """INSERT INTO credit_lines (company_id, facility_id, product_name, credit_type,
             credit_amount, used_amount, interest_rate, guarantee_type,
             start_date, end_date, status, notes)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
        (company_id, body.facility_id, body.product_name, body.credit_type,
         body.credit_amount, body.used_amount, body.interest_rate, body.guarantee_type,
         body.start_date, body.end_date, body.status, body.notes)
    )
    conn.commit()
    row = conn.execute("SELECT * FROM credit_lines WHERE id=?", (cur.lastrowid,)).fetchone()
    conn.close()
    return dict(row)

@app.put("/api/credit-lines/{line_id}")
def update_credit_line(line_id: int, body: CreditLineBody):
    conn = get_db()
    conn.execute(
        """UPDATE credit_lines SET facility_id=?, product_name=?, credit_type=?,
             credit_amount=?, used_amount=?, interest_rate=?, guarantee_type=?,
             start_date=?, end_date=?, status=?, notes=?
           WHERE id=?""",
        (body.facility_id, body.product_name, body.credit_type,
         body.credit_amount, body.used_amount, body.interest_rate, body.guarantee_type,
         body.start_date, body.end_date, body.status, body.notes, line_id)
    )
    conn.commit()
    row = conn.execute("SELECT * FROM credit_lines WHERE id=?", (line_id,)).fetchone()
    conn.close()
    return dict(row)

@app.delete("/api/credit-lines/{line_id}")
def delete_credit_line(line_id: int):
    conn = get_db()
    conn.execute("DELETE FROM credit_lines WHERE id=?", (line_id,))
    conn.commit()
    conn.close()
    return {"ok": True}

@app.get("/api/credit-lines/expiring")
def expiring_credit_lines(days: int = 30):
    from datetime import date, timedelta
    today = date.today().isoformat()
    deadline = (date.today() + timedelta(days=days)).isoformat()
    conn = get_db()
    rows = conn.execute(
        """SELECT cl.*, c.name as company_name FROM credit_lines cl
           JOIN companies c ON cl.company_id=c.id
           WHERE cl.end_date != '' AND cl.end_date BETWEEN ? AND ? AND cl.status != '已结清'
           ORDER BY cl.end_date""",
        (today, deadline)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


CREDIT_EXTRACT_PROMPT = """你是银行信贷助手，负责从授信批复文件中提取结构化信息。

请从以下文本中提取授信批复的完整内容，严格按JSON格式返回，不加任何说明：

{
  "approval_no": "批复文号，如××银发[2024]001号",
  "max_amount": 最高控制额度数字（万元，纯数字），
  "start_date": "生效日期YYYY-MM-DD",
  "end_date": "到期日期YYYY-MM-DD",
  "facilities": [
    {
      "facility_type": "组合额度 或 特别授信额度 或 专项额度",
      "name": "额度名称（如有，否则空字符串）",
      "approved_amount": 额度金额（万元，纯数字）,
      "products": [
        {
          "product_name": "业务名称",
          "credit_type": "业务品种，如流动资金贷款/固定资产贷款/承兑汇票/信用证/保函/贸易融资",
          "credit_amount": 业务金额（万元，纯数字），
          "guarantee_type": "担保方式，如信用/抵押/质押/保证/抵押+保证",
          "interest_rate": "利率表述，如LPR+50BP或4.35%"
        }
      ]
    }
  ]
}

注意：
- 若文中没有明确区分子额度类型，所有业务放入一个"组合额度"下
- 金额统一转为万元，纯数字
- 日期统一为YYYY-MM-DD格式
- 只返回JSON，不加任何说明和markdown代码块

文档内容：
"""

@app.post("/api/companies/{company_id}/credit-facilities/extract-pdf")
async def extract_credit_pdf(company_id: int, file: UploadFile = File(...),
                              approval_type: str = "原始授信",
                              parent_approval_id: str = ""):
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(400, "请上传PDF文件")
    content = await file.read()
    tmp_path = UPLOAD_DIR / f"_tmp_credit_{uuid.uuid4().hex}.pdf"
    tmp_path.write_bytes(content)
    try:
        text = await _extract_text_hybrid_pdf(tmp_path, max_pages=30)
        if not text.strip():
            raise HTTPException(400, "PDF无法提取文字，请确认文件完整")

        api_key = os.getenv("DEEPSEEK_API_KEY", "")
        base_url = os.getenv("DEEPSEEK_BASE_URL", "")
        if not api_key:
            raise HTTPException(400, "未配置AI接口，无法解析")

        import re
        reply = chat([{"role": "user", "content": CREDIT_EXTRACT_PROMPT + text[:8000]}], mode="cloud")
        # 清理可能的代码块标记
        reply = re.sub(r"```(?:json)?", "", reply).strip().strip("`").strip()
        data = json.loads(reply)
        data["_approval_type"] = approval_type
        data["_parent_approval_id"] = int(parent_approval_id) if parent_approval_id.isdigit() else None
        return data
    except json.JSONDecodeError:
        raise HTTPException(500, "AI返回格式异常，请重试或手动录入")
    except HTTPException:
        raise
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(500, f"提取失败：{e}")
    finally:
        tmp_path.unlink(missing_ok=True)


# ── 贷后检查 ──────────────────────────────────────────────────────────────────

class PostLoanCheckBody(BaseModel):
    check_date: str
    check_type: str = "日常检查"
    risk_level: str = "正常"
    inspector: str = ""
    content: str = ""
    issues: str = ""
    measures: str = ""
    next_check_date: str = ""

@app.get("/api/companies/{company_id}/post-loan-checks")
def list_post_loan_checks(company_id: int):
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM post_loan_checks WHERE company_id=? ORDER BY check_date DESC",
        (company_id,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/api/companies/{company_id}/post-loan-checks")
def create_post_loan_check(company_id: int, body: PostLoanCheckBody):
    conn = get_db()
    cur = conn.execute(
        """INSERT INTO post_loan_checks (company_id, check_date, check_type, risk_level,
             inspector, content, issues, measures, next_check_date)
           VALUES (?,?,?,?,?,?,?,?,?)""",
        (company_id, body.check_date, body.check_type, body.risk_level,
         body.inspector, body.content, body.issues, body.measures, body.next_check_date)
    )
    conn.commit()
    row = conn.execute("SELECT * FROM post_loan_checks WHERE id=?", (cur.lastrowid,)).fetchone()
    conn.close()
    return dict(row)

@app.put("/api/post-loan-checks/{check_id}")
def update_post_loan_check(check_id: int, body: PostLoanCheckBody):
    conn = get_db()
    conn.execute(
        """UPDATE post_loan_checks SET check_date=?, check_type=?, risk_level=?,
             inspector=?, content=?, issues=?, measures=?, next_check_date=?
           WHERE id=?""",
        (body.check_date, body.check_type, body.risk_level,
         body.inspector, body.content, body.issues, body.measures, body.next_check_date,
         check_id)
    )
    conn.commit()
    row = conn.execute("SELECT * FROM post_loan_checks WHERE id=?", (check_id,)).fetchone()
    conn.close()
    return dict(row)

@app.delete("/api/post-loan-checks/{check_id}")
def delete_post_loan_check(check_id: int):
    conn = get_db()
    conn.execute("DELETE FROM post_loan_checks WHERE id=?", (check_id,))
    conn.commit()
    conn.close()
    return {"ok": True}


@app.delete("/api/companies/{company_id}")
def delete_company(company_id: int):
    conn = get_db()
    conn.execute("DELETE FROM companies WHERE id=?", (company_id,))
    conn.commit()
    conn.close()
    return {"ok": True}


# ── Contacts ──────────────────────────────────────────────────────────────────

class ContactBody(BaseModel):
    company_id: int
    name: str
    position: str = ""
    phone: str = ""
    wechat: str = ""
    email: str = ""
    is_primary: bool = False
    notes: str = ""


@app.get("/api/companies/{company_id}/contacts")
def list_contacts(company_id: int):
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM contacts WHERE company_id=? ORDER BY is_primary DESC, created_at",
        (company_id,),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.post("/api/contacts")
def create_contact(body: ContactBody):
    conn = get_db()
    if body.is_primary:
        conn.execute("UPDATE contacts SET is_primary=0 WHERE company_id=?", (body.company_id,))
    cur = conn.execute(
        """INSERT INTO contacts (company_id, name, position, phone, wechat, email, is_primary, notes)
           VALUES (?,?,?,?,?,?,?,?)""",
        (body.company_id, body.name, body.position, body.phone,
         body.wechat, body.email, 1 if body.is_primary else 0, body.notes),
    )
    conn.execute(
        "UPDATE companies SET updated_at=datetime('now','localtime') WHERE id=?",
        (body.company_id,),
    )
    conn.commit()
    row = conn.execute("SELECT * FROM contacts WHERE id=?", (cur.lastrowid,)).fetchone()
    conn.close()
    return dict(row)


@app.put("/api/contacts/{contact_id}")
def update_contact(contact_id: int, body: ContactBody):
    conn = get_db()
    if body.is_primary:
        conn.execute("UPDATE contacts SET is_primary=0 WHERE company_id=?", (body.company_id,))
    conn.execute(
        """UPDATE contacts SET name=?, position=?, phone=?, wechat=?, email=?, is_primary=?, notes=?
           WHERE id=?""",
        (body.name, body.position, body.phone, body.wechat, body.email,
         1 if body.is_primary else 0, body.notes, contact_id),
    )
    conn.commit()
    row = conn.execute("SELECT * FROM contacts WHERE id=?", (contact_id,)).fetchone()
    conn.close()
    return dict(row)


@app.delete("/api/contacts/{contact_id}")
def delete_contact(contact_id: int):
    conn = get_db()
    conn.execute("DELETE FROM contacts WHERE id=?", (contact_id,))
    conn.commit()
    conn.close()
    return {"ok": True}


# ── Interactions ──────────────────────────────────────────────────────────────

class InteractionBody(BaseModel):
    company_id: int
    contact_id: Optional[int] = None
    date: str
    type: str
    content: str
    next_action: str = ""


@app.get("/api/companies/{company_id}/interactions")
def list_interactions(company_id: int):
    conn = get_db()
    rows = conn.execute(
        """SELECT i.*, c.name as contact_name FROM interactions i
           LEFT JOIN contacts c ON i.contact_id=c.id
           WHERE i.company_id=? ORDER BY i.date DESC, i.created_at DESC""",
        (company_id,),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.post("/api/interactions")
def create_interaction(body: InteractionBody):
    conn = get_db()
    cur = conn.execute(
        "INSERT INTO interactions (company_id, contact_id, date, type, content, next_action) VALUES (?,?,?,?,?,?)",
        (body.company_id, body.contact_id, body.date, body.type, body.content, body.next_action),
    )
    conn.execute(
        "UPDATE companies SET updated_at=datetime('now','localtime') WHERE id=?",
        (body.company_id,),
    )
    conn.commit()
    row = conn.execute(
        """SELECT i.*, c.name as contact_name FROM interactions i
           LEFT JOIN contacts c ON i.contact_id=c.id WHERE i.id=?""",
        (cur.lastrowid,),
    ).fetchone()
    conn.close()
    return dict(row)


@app.delete("/api/interactions/{interaction_id}")
def delete_interaction(interaction_id: int):
    conn = get_db()
    conn.execute("DELETE FROM interactions WHERE id=?", (interaction_id,))
    conn.commit()
    conn.close()
    return {"ok": True}


# ── Todos ─────────────────────────────────────────────────────────────────────

class TodoBody(BaseModel):
    company_id: Optional[int] = None
    contact_id: Optional[int] = None
    date: str
    end_date: str = ""
    content: str
    priority: str = "medium"
    sub_items: list = []


@app.get("/api/todos")
def list_todos(date_filter: str = "", company_id: Optional[int] = None):
    conn = get_db()
    query = """
        SELECT t.*, co.name as company_name, c.name as contact_name,
               (SELECT COUNT(*) FROM todo_documents WHERE todo_id=t.id) as doc_count
        FROM todos t
        LEFT JOIN companies co ON t.company_id=co.id
        LEFT JOIN contacts c ON t.contact_id=c.id
        WHERE 1=1
    """
    params = []
    if date_filter:
        query += " AND t.date = ?"
        params.append(date_filter)
    if company_id:
        query += " AND t.company_id = ?"
        params.append(company_id)
    query += " ORDER BY t.done, t.priority DESC, t.date"
    rows = conn.execute(query, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.post("/api/todos")
def create_todo(body: TodoBody):
    conn = get_db()
    cur = conn.execute(
        "INSERT INTO todos (company_id, contact_id, date, end_date, content, priority, sub_items) VALUES (?,?,?,?,?,?,?)",
        (body.company_id, body.contact_id, body.date, body.end_date, body.content, body.priority,
         json.dumps(body.sub_items, ensure_ascii=False)),
    )
    conn.commit()
    row = conn.execute(
        """SELECT t.*, co.name as company_name, c.name as contact_name
           FROM todos t LEFT JOIN companies co ON t.company_id=co.id
           LEFT JOIN contacts c ON t.contact_id=c.id WHERE t.id=?""",
        (cur.lastrowid,),
    ).fetchone()
    conn.close()
    return dict(row)


@app.put("/api/todos/{todo_id}")
def update_todo(todo_id: int, body: TodoBody):
    conn = get_db()
    conn.execute(
        "UPDATE todos SET date=?, end_date=?, content=?, priority=?, sub_items=? WHERE id=?",
        (body.date, body.end_date, body.content, body.priority,
         json.dumps(body.sub_items, ensure_ascii=False), todo_id),
    )
    conn.commit()
    row = conn.execute(
        """SELECT t.*, co.name as company_name FROM todos t
           LEFT JOIN companies co ON t.company_id=co.id WHERE t.id=?""",
        (todo_id,),
    ).fetchone()
    conn.close()
    return dict(row)


class SubItemsBody(BaseModel):
    sub_items: list = []

@app.patch("/api/todos/{todo_id}/sub_items")
def update_sub_items(todo_id: int, body: SubItemsBody):
    conn = get_db()
    conn.execute("UPDATE todos SET sub_items=? WHERE id=?",
                 (json.dumps(body.sub_items, ensure_ascii=False), todo_id))
    conn.commit()
    conn.close()
    return {"ok": True}


@app.patch("/api/todos/{todo_id}/toggle")
def toggle_todo(todo_id: int):
    conn = get_db()
    conn.execute("UPDATE todos SET done = 1 - done WHERE id=?", (todo_id,))
    conn.commit()
    row = conn.execute(
        """SELECT t.*, co.name as company_name, c.name as contact_name
           FROM todos t LEFT JOIN companies co ON t.company_id=co.id
           LEFT JOIN contacts c ON t.contact_id=c.id WHERE t.id=?""",
        (todo_id,),
    ).fetchone()
    conn.close()
    return dict(row)


@app.delete("/api/todos/{todo_id}")
def delete_todo(todo_id: int):
    conn = get_db()
    conn.execute("DELETE FROM todos WHERE id=?", (todo_id,))
    conn.commit()
    conn.close()
    return {"ok": True}


# ── Todo Documents ─────────────────────────────────────────────────────────────

TODO_DOC_ALLOWED = {
    ".pdf", ".doc", ".docx", ".xls", ".xlsx",
    ".ppt", ".pptx", ".txt", ".csv", ".md",
    ".jpg", ".jpeg", ".png",
}

@app.get("/api/todos/{todo_id}/documents")
def list_todo_documents(todo_id: int):
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM todo_documents WHERE todo_id=? ORDER BY uploaded_at DESC",
        (todo_id,),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.post("/api/todos/{todo_id}/documents")
async def upload_todo_document(todo_id: int, file: UploadFile = File(...)):
    ext = Path(file.filename).suffix.lower()
    if ext not in TODO_DOC_ALLOWED:
        raise HTTPException(400, f"不支持的文件格式")
    saved_name = f"todo_{uuid.uuid4().hex}{ext}"
    dest = UPLOAD_DIR / saved_name
    content = await file.read()
    dest.write_bytes(content)
    conn = get_db()
    cur = conn.execute(
        "INSERT INTO todo_documents (todo_id, filename, original_name, size) VALUES (?,?,?,?)",
        (todo_id, saved_name, file.filename, len(content)),
    )
    conn.commit()
    row = conn.execute("SELECT * FROM todo_documents WHERE id=?", (cur.lastrowid,)).fetchone()
    conn.close()
    return dict(row)


@app.get("/api/todo_documents/{doc_id}/download")
def download_todo_document(doc_id: int):
    conn = get_db()
    row = conn.execute("SELECT * FROM todo_documents WHERE id=?", (doc_id,)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(404, "文件不存在")
    path = UPLOAD_DIR / row["filename"]
    if not path.exists():
        raise HTTPException(404, "文件已丢失")
    ext = Path(row["original_name"]).suffix.lower()
    mime = MIME_TYPES.get(ext, "application/octet-stream")
    if ext in {".jpg", ".jpeg"}:
        mime = "image/jpeg"
    elif ext == ".png":
        mime = "image/png"
    return FileResponse(path, filename=row["original_name"], media_type=mime)


@app.delete("/api/todo_documents/{doc_id}")
def delete_todo_document(doc_id: int):
    conn = get_db()
    row = conn.execute("SELECT * FROM todo_documents WHERE id=?", (doc_id,)).fetchone()
    if row:
        path = UPLOAD_DIR / row["filename"]
        if path.exists():
            path.unlink()
        conn.execute("DELETE FROM todo_documents WHERE id=?", (doc_id,))
        conn.commit()
    conn.close()
    return {"ok": True}


# ── AI ────────────────────────────────────────────────────────────────────────

class AIChatBody(BaseModel):
    message: str
    company_id: Optional[int] = None
    mode: str = "fast"
    history: list = []


@app.post("/api/ai/chat")
def ai_chat(body: AIChatBody):
    messages = list(body.history)

    if body.company_id:
        conn = get_db()
        company = row_to_dict(
            conn.execute("SELECT * FROM companies WHERE id=?", (body.company_id,)).fetchone()
        )
        contacts = [dict(r) for r in conn.execute(
            "SELECT * FROM contacts WHERE company_id=? ORDER BY is_primary DESC", (body.company_id,)
        ).fetchall()]
        interactions = [dict(r) for r in conn.execute(
            """SELECT i.*, c.name as contact_name FROM interactions i
               LEFT JOIN contacts c ON i.contact_id=c.id
               WHERE i.company_id=? ORDER BY i.date DESC LIMIT 30""",
            (body.company_id,),
        ).fetchall()]
        todos = [dict(r) for r in conn.execute(
            "SELECT * FROM todos WHERE company_id=? ORDER BY date", (body.company_id,)
        ).fetchall()]
        conn.close()

        context = build_company_context(company, contacts, interactions, todos)
        fin_ctx = _build_fin_context(body.company_id)
        full_ctx = context + ("\n\n" + fin_ctx if fin_ctx else "")
        messages.insert(0, {"role": "user", "content": f"以下是该企业的档案信息：\n\n{full_ctx}"})
        messages.insert(1, {"role": "assistant", "content": "好的，我已了解该企业的档案信息，请问您有什么需要分析或处理的？"})

        doc_chunks = _search_doc_chunks(body.company_id, body.message)
        if doc_chunks:
            doc_ctx = "\n\n".join(doc_chunks)
            messages.insert(2, {"role": "user", "content": f"以下是与当前问题相关的文档摘录：\n\n{doc_ctx}"})
            messages.insert(3, {"role": "assistant", "content": "好的，我已参考相关文档内容。"})

    messages.append({"role": "user", "content": body.message})

    try:
        if body.mode == "web":
            reply = chat_with_web_search(messages)
        else:
            reply = chat(messages, mode=body.mode)
        return {"reply": reply, "mode": body.mode}
    except Exception as e:
        raise HTTPException(500, f"AI 服务错误：{str(e)}")


@app.post("/api/ai/search")
def ai_search(body: AIChatBody):
    conn = get_db()
    companies = [row_to_dict(r) for r in conn.execute(
        "SELECT * FROM companies ORDER BY level, name"
    ).fetchall()]

    blocks = [f"=== 客户档案总览（共 {len(companies)} 家企业）==="]
    for co in companies:
        contacts = conn.execute(
            "SELECT name, position, phone FROM contacts WHERE company_id=? ORDER BY is_primary DESC LIMIT 3",
            (co["id"],),
        ).fetchall()
        recent = conn.execute(
            "SELECT date, type, content, next_action FROM interactions WHERE company_id=? ORDER BY date DESC LIMIT 5",
            (co["id"],),
        ).fetchall()
        todos = conn.execute(
            "SELECT date, content FROM todos WHERE company_id=? AND done=0 ORDER BY date LIMIT 3",
            (co["id"],),
        ).fetchall()

        products = "、".join(co.get("products") or []) or "—"
        block = [f"\n【{co['name']}】等级:{co['level']} | 行业:{co['industry'] or '—'} | 授信:{co['credit_limit']}万 | 产品:{products}"]
        if contacts:
            ct = " / ".join(f"{r[0]}{'·'+r[1] if r[1] else ''}{' '+r[2] if r[2] else ''}" for r in contacts)
            block.append(f"  联系人: {ct}")
        for r in recent:
            line = f"  [{r[0]}]{r[1]}: {r[2][:60]}"
            if r[3]:
                line += f" → {r[3][:30]}"
            block.append(line)
        for t in todos:
            block.append(f"  待办[{t[0]}]: {t[1][:40]}")
        blocks.extend(block)

    conn.close()
    context = "\n".join(blocks)

    messages = [
        {"role": "user", "content": f"以下是我管理的全部客户档案信息：\n\n{context}"},
        {"role": "assistant", "content": "好的，我已了解您所有客户的档案，请问您想查询什么？"},
    ]
    messages.extend(body.history)
    messages.append({"role": "user", "content": body.message})

    try:
        reply = chat(messages, mode="cloud")
        return {"reply": reply, "mode": "cloud"}
    except Exception as e:
        raise HTTPException(500, f"AI 服务错误：{str(e)}")


@app.get("/api/ai/company-research/{company_id}")
def ai_company_research(company_id: int):
    conn = get_db()
    row = conn.execute("SELECT name, industry FROM companies WHERE id=?", (company_id,)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(404, "企业不存在")

    name, industry = row[0], row[1] or "所在行业"
    prompt = f"""请联网搜索企业「{name}」的最新公开信息，从银行客户经理视角输出以下内容：

1. **企业概况**：成立时间、注册地、实控人/股权结构、注册资本
2. **主营业务**：核心产品或服务、主要客户/市场
3. **行业地位**：在{industry}中的竞争地位、规模排名
4. **财务概况**：营收量级、盈利状况（近年公开数据或估算）
5. **信用与合规风险**：公开失信记录、行政处罚、重大诉讼纠纷
6. **近期重要动态**：融资、重大合同、人事变动、扩张或收缩迹象
7. **银行合作视角**：主要授信机遇与潜在风险点

请优先使用最新网络数据。若该企业规模较小信息有限，请如实说明。"""

    try:
        reply = chat_with_web_search([{"role": "user", "content": prompt}])
        return {"reply": reply}
    except Exception as e:
        raise HTTPException(500, f"AI 服务错误：{str(e)}")


@app.get("/api/ai/company-autofill/{company_id}")
def ai_company_autofill(company_id: int):
    conn = get_db()
    row = conn.execute("SELECT name FROM companies WHERE id=?", (company_id,)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(404, "企业不存在")

    name = row[0]
    # Step 1: Kimi searches freely in natural language (no JSON constraint)
    search_prompt = f"""请联网搜索「{name}」的详细企业信息，需要找到以下内容：
1. 法定代表人/董事长姓名
2. 统一社会信用代码（18位）
3. 注册资本
4. 成立日期
5. 工商注册地址
6. 实际办公地址
7. 主营业务简介
8. 工商登记经营范围（完整文本）
9. 所属行业
10. 企业规模（大型/中型/小型/微型）
11. 员工人数

请尽量从天眼查、企查查、官网等多个来源核实，给出准确信息。"""

    try:
        search_result = chat_with_web_search([{"role": "user", "content": search_prompt}], max_tokens=4096)

        # Step 2: local fast model extracts structured JSON from the search result
        extract_prompt = f"""从下面的企业信息文本中提取结构化数据，只返回JSON不要其他内容，找不到的字段填空字符串""：

{search_result}

返回格式：
{{"legal_rep":"法定代表人姓名","credit_code":"18位社会信用代码","reg_capital":"注册资本","established_date":"成立日期YYYY-MM-DD","reg_address":"注册地址","biz_scope":"主营业务50字内","industry":"行业10字内","company_scale":"大型企业或中型企业或小型企业或微型企业","office_address":"办公地址","employee_count":"员工人数","operating_scope":"工商登记经营范围完整文本"}}"""

        import re as _re
        raw = chat([{"role": "user", "content": extract_prompt}], mode="fast")
        raw_clean = _re.sub(r'```(?:json)?\s*', '', raw).strip().rstrip('`').strip()
        start, end = raw_clean.find('{'), raw_clean.rfind('}')
        fields = json.loads(raw_clean[start:end + 1]) if start != -1 else {}
        placeholder_words = {"法定代表人姓名", "统一社会信用代码", "注册资本", "注册地址", "主营业务", "所属行业", "成立日期", "员工人数", "办公地址"}
        scale_map = {"大型": "大型企业", "中型": "中型企业", "小型": "小型企业", "微型": "微型企业"}
        cleaned = {}
        for k, v in fields.items():
            v = (v or "").strip()
            if not v or v in placeholder_words or "（" in v[:6]:
                continue
            if k == "company_scale":
                v = scale_map.get(v, v)
                if v not in {"大型企业", "中型企业", "小型企业", "微型企业"}:
                    continue
            cleaned[k] = v[:100] if k == "biz_scope" else v
        return {"fields": cleaned}
    except Exception as e:
        raise HTTPException(500, f"AI 服务错误：{str(e)}")


class ExtractBody(BaseModel):
    text: str

@app.post("/api/ai/extract-fields")
def ai_extract_fields(body: ExtractBody):
    import re as _re
    extract_prompt = f"""从下面的企业信息文本中提取结构化数据，只返回JSON不要其他内容，找不到的字段填空字符串""：

{body.text}

返回格式：
{{"legal_rep":"法定代表人姓名","credit_code":"18位社会信用代码","reg_capital":"注册资本","established_date":"成立日期YYYY-MM-DD","reg_address":"注册地址","biz_scope":"主营业务50字内","industry":"行业10字内","company_scale":"大型企业或中型企业或小型企业或微型企业","office_address":"办公地址","employee_count":"员工人数","operating_scope":"工商登记经营范围完整文本"}}"""
    try:
        raw = chat([{"role": "user", "content": extract_prompt}], mode="fast")
        raw_clean = _re.sub(r'```(?:json)?\s*', '', raw).strip().rstrip('`').strip()
        start, end = raw_clean.find('{'), raw_clean.rfind('}')
        fields = json.loads(raw_clean[start:end + 1]) if start != -1 else {}
        placeholder_words = {"法定代表人姓名", "统一社会信用代码", "注册资本", "注册地址", "主营业务", "所属行业", "成立日期", "员工人数", "办公地址"}
        scale_map = {"大型": "大型企业", "中型": "中型企业", "小型": "小型企业", "微型": "微型企业"}
        cleaned = {}
        for k, v in fields.items():
            v = (v or "").strip()
            if not v or v in placeholder_words or "（" in v[:6]:
                continue
            if k == "company_scale":
                v = scale_map.get(v, v)
                if v not in {"大型企业", "中型企业", "小型企业", "微型企业"}:
                    continue
            cleaned[k] = v[:200] if k == "operating_scope" else (v[:100] if k == "biz_scope" else v)
        return {"fields": cleaned}
    except Exception as e:
        raise HTTPException(500, f"AI 服务错误：{str(e)}")


@app.get("/api/ai/daily-brief")
def daily_brief():
    today = date.today().isoformat()
    from datetime import timedelta
    deadline_30 = (date.today() + timedelta(days=30)).isoformat()
    conn = get_db()

    todos = conn.execute(
        """SELECT t.*, co.name as company_name
           FROM todos t LEFT JOIN companies co ON t.company_id=co.id
           WHERE t.date <= ? AND t.done=0 ORDER BY t.priority DESC""",
        (today,),
    ).fetchall()

    expiring_credits = conn.execute(
        """SELECT cl.*, c.name as company_name FROM credit_lines cl
           JOIN companies c ON cl.company_id=c.id
           WHERE cl.end_date != '' AND cl.end_date BETWEEN ? AND ? AND cl.status != '已结清'
           ORDER BY cl.end_date LIMIT 10""",
        (today, deadline_30)
    ).fetchall()

    overdue_checks = conn.execute(
        """SELECT c.name as company_name, c.id as company_id,
               MAX(plc.check_date) as last_check,
               MIN(plc.next_check_date) as next_due
           FROM companies c
           JOIN post_loan_checks plc ON plc.company_id=c.id
           WHERE c.business_stage IN ('用信中','贷后管理')
             AND plc.next_check_date != '' AND plc.next_check_date <= ?
           GROUP BY c.id ORDER BY next_due LIMIT 10""",
        (today,)
    ).fetchall()

    conn.close()

    parts = []
    if todos:
        todo_text = "\n".join(
            f"- {'[' + r['company_name'] + '] ' if r['company_name'] else ''}[{r['date']}] {r['content']}"
            for r in todos
        )
        parts.append(f"【待办事项（{len(todos)}条）】\n{todo_text}")

    if expiring_credits:
        cl_text = "\n".join(
            f"- [{r['company_name']}] {r['product_name']} 到期{r['end_date']}，授信{r['credit_amount']}万元"
            for r in expiring_credits
        )
        parts.append(f"【30天内到期授信（{len(expiring_credits)}条）】\n{cl_text}")

    if overdue_checks:
        chk_text = "\n".join(
            f"- [{r['company_name']}] 贷后检查应于{r['next_due']}完成"
            for r in overdue_checks
        )
        parts.append(f"【逾期贷后检查（{len(overdue_checks)}家）】\n{chk_text}")

    if not parts:
        return {"reply": "今日没有待办事项，授信和贷后检查均无逾期，可以主动拓展新客户。", "mode": "fast"}

    prompt = (f"今日是{today}，我是银行对公客户经理，以下是今日工作摘要：\n\n"
              + "\n\n".join(parts)
              + "\n\n请帮我梳理今日工作优先级，分类给出行动建议，要简洁实用，突出最紧急的事项。")

    try:
        reply = chat([{"role": "user", "content": prompt}], mode="fast")
        return {"reply": reply, "mode": "fast"}
    except Exception as e:
        return {"reply": f"AI 暂时不可用：{str(e)}", "mode": "error"}


# ── PDF Import ────────────────────────────────────────────────────────────────

PDF_PARSE_PROMPT = """你是数据提取助手。从以下PDF文本中提取所有企业客户信息，返回JSON数组。

每个对象代表一家企业，包含：
- name: 企业名称
- industry: 行业
- level: 客户等级（A/B/C，不确定填B）
- credit_limit: 授信额度，纯数字，单位万元（没有填0）
- products: 合作产品数组，如["贷款","存款"]
- notes: 备注
- contacts: 联系人数组，每个联系人包含 {name, position, phone, wechat, email, is_primary}

只返回JSON数组，不要任何解释。格式：
[{"name":"XX公司","industry":"","level":"B","credit_limit":0,"products":[],"notes":"","contacts":[{"name":"张总","position":"董事长","phone":"138...","wechat":"","email":"","is_primary":true}]}]

PDF文本：
"""


@app.post("/api/import/pdf/preview")
async def pdf_preview(file: UploadFile = File(...)):
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(400, "请上传 PDF 文件")

    import pdfplumber, io
    content = await file.read()
    text_parts = []
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                text_parts.append(t)

    full_text = "\n".join(text_parts).strip()
    if not full_text:
        raise HTTPException(400, "PDF 中未提取到文字，请确认已 OCR 处理")

    if len(full_text) > 12000:
        full_text = full_text[:12000] + "\n...(内容过长已截断)"

    try:
        reply = chat([{"role": "user", "content": PDF_PARSE_PROMPT + full_text}], mode="fast")
        start = reply.find("[")
        end = reply.rfind("]") + 1
        if start == -1 or end == 0:
            raise ValueError("AI 未返回有效 JSON")
        companies = json.loads(reply[start:end])
        return {"companies": companies, "total": len(companies)}
    except Exception as e:
        raise HTTPException(500, f"AI 解析失败：{str(e)}")


@app.post("/api/import/pdf/confirm")
async def pdf_confirm(request: Request):
    body = await request.json()
    companies_data = body.get("companies", [])
    if not companies_data:
        raise HTTPException(400, "没有要导入的数据")

    conn = get_db()
    imported = 0
    skipped = 0
    for c in companies_data:
        if not c.get("name"):
            skipped += 1
            continue
        cur = conn.execute(
            """INSERT INTO companies (name, industry, level, credit_limit, products, tags, notes)
               VALUES (?,?,?,?,?,?,?)""",
            (c.get("name", ""), c.get("industry", ""), c.get("level", "B"),
             float(c.get("credit_limit", 0) or 0),
             json.dumps(c.get("products", []), ensure_ascii=False), "[]",
             c.get("notes", "")),
        )
        company_id = cur.lastrowid
        for idx, ct in enumerate(c.get("contacts", [])):
            if not ct.get("name"):
                continue
            conn.execute(
                """INSERT INTO contacts (company_id, name, position, phone, wechat, email, is_primary)
                   VALUES (?,?,?,?,?,?,?)""",
                (company_id, ct.get("name", ""), ct.get("position", ""),
                 ct.get("phone", ""), ct.get("wechat", ""), ct.get("email", ""),
                 1 if (ct.get("is_primary") or idx == 0) else 0),
            )
        imported += 1
    conn.commit()
    conn.close()
    return {"imported": imported, "skipped": skipped}


# ── Documents ─────────────────────────────────────────────────────────────────

UPLOAD_DIR = Path("uploads")

ALLOWED_EXTENSIONS = {
    ".pdf", ".doc", ".docx", ".xls", ".xlsx",
    ".ppt", ".pptx", ".txt", ".csv", ".md",
}

MIME_TYPES = {
    ".pdf":  "application/pdf",
    ".doc":  "application/msword",
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ".xls":  "application/vnd.ms-excel",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".ppt":  "application/vnd.ms-powerpoint",
    ".pptx": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
    ".txt":  "text/plain",
    ".csv":  "text/csv",
    ".md":   "text/markdown",
}


@app.get("/api/companies/{company_id}/documents")
def list_documents(company_id: int):
    conn = get_db()
    rows = conn.execute(
        """SELECT id, company_id, filename, original_name, size, category,
                  description, uploaded_at, doc_indexed,
                  CASE WHEN doc_text != '' THEN 1 ELSE 0 END as doc_text
           FROM documents WHERE company_id=? ORDER BY uploaded_at DESC""",
        (company_id,),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.post("/api/companies/{company_id}/documents")
async def upload_document(
    company_id: int,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    category: str = Form("其他"),
):
    ext = Path(file.filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(400, f"不支持的文件格式，支持：{', '.join(ALLOWED_EXTENSIONS)}")
    saved_name = f"{uuid.uuid4().hex}{ext}"
    dest = UPLOAD_DIR / saved_name
    content = await file.read()
    dest.write_bytes(content)
    conn = get_db()
    cur = conn.execute(
        "INSERT INTO documents (company_id, filename, original_name, size, category, doc_indexed) VALUES (?,?,?,?,?,0)",
        (company_id, saved_name, file.filename, len(content), category),
    )
    conn.commit()
    doc_id = cur.lastrowid
    row = conn.execute("SELECT * FROM documents WHERE id=?", (doc_id,)).fetchone()
    conn.close()
    background_tasks.add_task(_index_document, doc_id, dest, file.filename)
    return dict(row)


@app.get("/api/documents/{doc_id}/download")
def download_document(doc_id: int):
    conn = get_db()
    row = conn.execute("SELECT * FROM documents WHERE id=?", (doc_id,)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(404, "文件不存在")
    path = UPLOAD_DIR / row["filename"]
    if not path.exists():
        raise HTTPException(404, "文件已丢失")
    ext = Path(row["original_name"]).suffix.lower()
    mime = MIME_TYPES.get(ext, "application/octet-stream")
    return FileResponse(path, filename=row["original_name"], media_type=mime)


@app.delete("/api/documents/{doc_id}")
def delete_document(doc_id: int):
    conn = get_db()
    row = conn.execute("SELECT * FROM documents WHERE id=?", (doc_id,)).fetchone()
    if row:
        path = UPLOAD_DIR / row["filename"]
        if path.exists():
            path.unlink()
        conn.execute("DELETE FROM documents WHERE id=?", (doc_id,))
        conn.execute("DELETE FROM documents_fts WHERE rowid=?", (doc_id,))
        conn.commit()
    conn.close()
    return {"ok": True}


# ── Bark Push ─────────────────────────────────────────────────────────────────

@app.get("/api/push/vapid-key")
def push_vapid_key():
    return {"public_key": "bark", "enabled": is_bark_enabled()}

@app.post("/api/push/subscribe")
def push_subscribe():
    enable_bark()
    return {"ok": True}

@app.delete("/api/push/subscribe")
def push_unsubscribe():
    disable_bark()
    return {"ok": True}

@app.post("/api/push/test")
def push_test():
    if not is_bark_enabled():
        raise HTTPException(400, "Bark 推送未开启，请先点击「开启推送」")
    ok = send_bark_push("📋 推送测试成功", "每天早上 8 点将自动发送今日待办提醒")
    if not ok:
        raise HTTPException(500, "推送发送失败，请检查服务器网络")
    return {"ok": True}


# ── 财务状况 ─────────────────────────────────────────────────────────────────

@app.get("/api/companies/{company_id}/financials")
def list_financials(company_id: int):
    conn = get_db()
    rows = conn.execute(
        "SELECT id, year, unit, source_doc, extracted_at FROM company_financials WHERE company_id=? ORDER BY year DESC",
        (company_id,),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.get("/api/companies/{company_id}/financials/{year}")
def get_financial(company_id: int, year: int):
    conn = get_db()
    row = conn.execute(
        "SELECT * FROM company_financials WHERE company_id=? AND year=?",
        (company_id, year),
    ).fetchone()
    conn.close()
    if not row:
        raise HTTPException(404, "无此年份数据")
    d = dict(row)
    for field in ("balance_sheet", "income", "balance_sheet_parent", "income_parent",
                  "cash_flow_consolidated", "cash_flow_parent"):
        d[field] = json.loads(d.get(field) or "[]")
    return d


class FinancialBody(BaseModel):
    year: int
    unit: str = "万元"
    balance_sheet: list = []
    income: list = []
    balance_sheet_parent: list = []
    income_parent: list = []
    cash_flow_consolidated: list = []
    cash_flow_parent: list = []
    source_doc: str = ""


@app.post("/api/companies/{company_id}/financials")
def save_financial(company_id: int, body: FinancialBody):
    conn = get_db()
    conn.execute(
        """INSERT INTO company_financials
             (company_id, year, unit, balance_sheet, income,
              balance_sheet_parent, income_parent, cash_flow_consolidated, cash_flow_parent, source_doc)
           VALUES (?,?,?,?,?,?,?,?,?,?)
           ON CONFLICT(company_id, year) DO UPDATE SET
             unit=excluded.unit,
             balance_sheet=excluded.balance_sheet,
             income=excluded.income,
             balance_sheet_parent=excluded.balance_sheet_parent,
             income_parent=excluded.income_parent,
             cash_flow_consolidated=excluded.cash_flow_consolidated,
             cash_flow_parent=excluded.cash_flow_parent,
             source_doc=excluded.source_doc,
             extracted_at=datetime('now','localtime')""",
        (company_id, body.year, body.unit,
         json.dumps(body.balance_sheet, ensure_ascii=False),
         json.dumps(body.income, ensure_ascii=False),
         json.dumps(body.balance_sheet_parent, ensure_ascii=False),
         json.dumps(body.income_parent, ensure_ascii=False),
         json.dumps(body.cash_flow_consolidated, ensure_ascii=False),
         json.dumps(body.cash_flow_parent, ensure_ascii=False),
         body.source_doc),
    )
    conn.commit()
    conn.close()
    return {"ok": True}


@app.delete("/api/companies/{company_id}/financials/{year}")
def delete_financial(company_id: int, year: int):
    conn = get_db()
    conn.execute("DELETE FROM company_financials WHERE company_id=? AND year=?", (company_id, year))
    conn.commit()
    conn.close()
    return {"ok": True}


VISION_MODEL = "moonshot-v1-8k-vision-preview"

_FIN_PROMPT = """这是审计报告的一页，请识别并提取财务数据，只返回JSON，不要任何说明文字。

【第一步】识别报表主体（scope）：
- 页面标题含"合并" → scope="consolidated"
- 页面标题含"母公司"或"本部" → scope="parent"
- 无明确标注 → scope="consolidated"

【第二步】识别报表类型（stmt）：
- 含"货币资金"/"应收账款"/"固定资产"/"资产总计"/"负债合计" → stmt="balance_sheet"
- 含"营业收入"/"营业成本"/"净利润"/"归属于母公司股东" → stmt="income"
- 含"经营活动产生"/"投资活动产生"/"筹资活动产生"/"现金及现金等价物" → stmt="cash_flow"
- 不属于以上任何报表 → stmt="other"

type字段 = stmt + "_" + scope（当stmt为"other"时type="other"）
示例："balance_sheet_consolidated" / "income_parent" / "cash_flow_consolidated"

JSON格式：
{"type":"...","year":年份整数,"prev_year":上年整数,"unit":"元或万元","items":[{"name":"科目","current":数字或null,"prev":数字或null,"type":"total/subtotal/header/detail"}]}

规则：
- year/prev_year：从表头"xxxx年12月31日"/"xxxx年度"识别报告期年份；扫描误读时从数值量级推断
- unit：从表头"单位：元"/"金额单位：元"/"单位：万元"识别
- item.type：total=资产总计/净利润/现金净增加额等最终汇总行，subtotal=小计/合计，header=流动资产/经营活动等章节标题（无数值），detail=明细科目
- 金额去千分位逗号，负数加"-"
- 不要遗漏任何有数值的行"""


def _to_wan(value, unit: str):
    if value is None:
        return None
    try:
        v = float(str(value).replace(",", "").replace("，", "").strip())
    except (ValueError, TypeError):
        return None
    if unit == "元":
        v = v / 10000
    elif unit == "亿元":
        v = v * 10000
    return round(v, 2)


# ── 标准科目模板 ──────────────────────────────────────────────────────────────
def _r(name, t):
    return {"name": name, "type": t}

_H, _D, _S, _T = "header", "detail", "subtotal", "total"

_BS_TMPL_COMMON_ASSETS = [
    _r("流动资产",                    _H),
    _r("货币资金",                    _D),
    _r("交易性金融资产",               _D),
    _r("衍生金融资产",                 _D),
    _r("应收票据",                    _D),
    _r("应收账款",                    _D),
    _r("应收款项融资",                 _D),
    _r("预付款项",                    _D),
    _r("其他应收款",                   _D),
    _r("存货",                        _D),
    _r("合同资产",                    _D),
    _r("持有待售资产",                 _D),
    _r("一年内到期的非流动资产",        _D),
    _r("其他流动资产",                 _D),
    _r("流动资产合计",                 _S),
    _r("非流动资产",                   _H),
    _r("债权投资",                    _D),
    _r("其他债权投资",                 _D),
    _r("长期应收款",                   _D),
    _r("长期股权投资",                 _D),
    _r("其他权益工具投资",              _D),
    _r("其他非流动金融资产",            _D),
    _r("投资性房地产",                 _D),
    _r("固定资产",                    _D),
    _r("在建工程",                    _D),
    _r("使用权资产",                   _D),
    _r("无形资产",                    _D),
    _r("开发支出",                    _D),
    _r("商誉",                        _D),
    _r("长期待摊费用",                 _D),
    _r("递延所得税资产",               _D),
    _r("其他非流动资产",               _D),
    _r("非流动资产合计",               _S),
    _r("资产总计",                    _T),
]

_BS_TMPL_COMMON_LIAB = [
    _r("流动负债",                    _H),
    _r("短期借款",                    _D),
    _r("交易性金融负债",               _D),
    _r("衍生金融负债",                 _D),
    _r("应付票据",                    _D),
    _r("应付账款",                    _D),
    _r("预收款项",                    _D),
    _r("合同负债",                    _D),
    _r("应付职工薪酬",                 _D),
    _r("应交税费",                    _D),
    _r("其他应付款",                   _D),
    _r("持有待售负债",                 _D),
    _r("一年内到期的非流动负债",        _D),
    _r("其他流动负债",                 _D),
    _r("流动负债合计",                 _S),
    _r("非流动负债",                   _H),
    _r("长期借款",                    _D),
    _r("应付债券",                    _D),
    _r("租赁负债",                    _D),
    _r("长期应付款",                   _D),
    _r("预计负债",                    _D),
    _r("递延收益",                    _D),
    _r("递延所得税负债",               _D),
    _r("其他非流动负债",               _D),
    _r("非流动负债合计",               _S),
    _r("负债合计",                    _T),
]

BS_CONSOL_TMPL = _BS_TMPL_COMMON_ASSETS + _BS_TMPL_COMMON_LIAB + [
    _r("所有者权益",                   _H),
    _r("股本",                        _D),
    _r("其他权益工具",                 _D),
    _r("资本公积",                    _D),
    _r("减：库存股",                   _D),
    _r("其他综合收益",                 _D),
    _r("专项储备",                    _D),
    _r("盈余公积",                    _D),
    _r("未分配利润",                   _D),
    _r("归属于母公司所有者权益合计",    _S),
    _r("少数股东权益",                 _D),
    _r("所有者权益合计",               _T),
    _r("负债和所有者权益总计",          _T),
]

BS_PARENT_TMPL = _BS_TMPL_COMMON_ASSETS + _BS_TMPL_COMMON_LIAB + [
    _r("所有者权益",                   _H),
    _r("股本",                        _D),
    _r("其他权益工具",                 _D),
    _r("资本公积",                    _D),
    _r("减：库存股",                   _D),
    _r("其他综合收益",                 _D),
    _r("专项储备",                    _D),
    _r("盈余公积",                    _D),
    _r("未分配利润",                   _D),
    _r("所有者权益合计",               _T),
    _r("负债和所有者权益总计",          _T),
]

INC_CONSOL_TMPL = [
    _r("营业收入",                    _T),
    _r("营业成本",                    _D),
    _r("税金及附加",                   _D),
    _r("销售费用",                    _D),
    _r("管理费用",                    _D),
    _r("研发费用",                    _D),
    _r("财务费用",                    _D),
    _r("其他收益",                    _D),
    _r("投资收益",                    _D),
    _r("公允价值变动收益",              _D),
    _r("信用减值损失",                 _D),
    _r("资产减值损失",                 _D),
    _r("资产处置收益",                 _D),
    _r("营业利润",                    _S),
    _r("营业外收入",                   _D),
    _r("营业外支出",                   _D),
    _r("利润总额",                    _S),
    _r("所得税费用",                   _D),
    _r("净利润",                      _T),
    _r("归属于母公司所有者的净利润",    _D),
    _r("少数股东损益",                 _D),
    _r("其他综合收益的税后净额",        _S),
    _r("综合收益总额",                 _T),
    _r("归属于母公司所有者的综合收益总额", _D),
    _r("归属于少数股东的综合收益总额",  _D),
    _r("基本每股收益",                 _D),
    _r("稀释每股收益",                 _D),
]

INC_PARENT_TMPL = [
    _r("营业收入",                    _T),
    _r("营业成本",                    _D),
    _r("税金及附加",                   _D),
    _r("销售费用",                    _D),
    _r("管理费用",                    _D),
    _r("研发费用",                    _D),
    _r("财务费用",                    _D),
    _r("其他收益",                    _D),
    _r("投资收益",                    _D),
    _r("公允价值变动收益",              _D),
    _r("信用减值损失",                 _D),
    _r("资产减值损失",                 _D),
    _r("资产处置收益",                 _D),
    _r("营业利润",                    _S),
    _r("营业外收入",                   _D),
    _r("营业外支出",                   _D),
    _r("利润总额",                    _S),
    _r("所得税费用",                   _D),
    _r("净利润",                      _T),
    _r("其他综合收益的税后净额",        _S),
    _r("综合收益总额",                 _T),
    _r("基本每股收益",                 _D),
    _r("稀释每股收益",                 _D),
]

CF_TMPL = [
    _r("经营活动产生的现金流量",                       _H),
    _r("销售商品、提供劳务收到的现金",                  _D),
    _r("收到的税费返还",                               _D),
    _r("收到其他与经营活动有关的现金",                  _D),
    _r("经营活动现金流入小计",                          _S),
    _r("购买商品、接受劳务支付的现金",                  _D),
    _r("支付给职工以及为职工支付的现金",                 _D),
    _r("支付的各项税费",                               _D),
    _r("支付其他与经营活动有关的现金",                  _D),
    _r("经营活动现金流出小计",                          _S),
    _r("经营活动产生的现金流量净额",                    _T),
    _r("投资活动产生的现金流量",                        _H),
    _r("收回投资收到的现金",                            _D),
    _r("取得投资收益收到的现金",                        _D),
    _r("处置固定资产等收回的现金净额",                  _D),
    _r("处置子公司及其他营业单位收到的现金净额",         _D),
    _r("收到其他与投资活动有关的现金",                  _D),
    _r("投资活动现金流入小计",                          _S),
    _r("购建固定资产、无形资产支付的现金",               _D),
    _r("投资支付的现金",                               _D),
    _r("取得子公司及其他营业单位支付的现金净额",         _D),
    _r("支付其他与投资活动有关的现金",                  _D),
    _r("投资活动现金流出小计",                          _S),
    _r("投资活动产生的现金流量净额",                    _T),
    _r("筹资活动产生的现金流量",                        _H),
    _r("吸收投资收到的现金",                            _D),
    _r("取得借款收到的现金",                            _D),
    _r("收到其他与筹资活动有关的现金",                  _D),
    _r("筹资活动现金流入小计",                          _S),
    _r("偿还债务支付的现金",                            _D),
    _r("分配股利、利润或偿付利息支付的现金",             _D),
    _r("支付其他与筹资活动有关的现金",                  _D),
    _r("筹资活动现金流出小计",                          _S),
    _r("筹资活动产生的现金流量净额",                    _T),
    _r("汇率变动对现金及现金等价物的影响",               _D),
    _r("现金及现金等价物净增加额",                      _T),
    _r("期初现金及现金等价物余额",                      _D),
    _r("期末现金及现金等价物余额",                      _T),
]

# 科目名称常见变体 → 标准名称
_NAME_ALIASES = {
    "资产合计": "资产总计",
    "负债和所有者权益合计": "负债和所有者权益总计",
    "负债及所有者权益总计": "负债和所有者权益总计",
    "归属于母公司股东的净利润": "归属于母公司所有者的净利润",
    "归属于母公司股东的综合收益总额": "归属于母公司所有者的综合收益总额",
    "归属于少数股东的净利润": "少数股东损益",
    "经营活动净现金流量": "经营活动产生的现金流量净额",
    "投资活动净现金流量": "投资活动产生的现金流量净额",
    "筹资活动净现金流量": "筹资活动产生的现金流量净额",
    "处置固定资产、无形资产和其他长期资产收回的现金净额": "处置固定资产等收回的现金净额",
    "购建固定资产、无形资产和其他长期资产支付的现金": "购建固定资产、无形资产支付的现金",
    "母公司所有者权益合计": "所有者权益合计",
}

import re as _re_fin


def _norm(name: str) -> str:
    """去掉编号前缀（一、二. 1. （一）等）和后缀括号注释，归一化空格"""
    n = _re_fin.sub(r'^[\s]*[（(]?[一二三四五六七八九十百\d]+[）)]?[、。.：: ]+', '', name)
    n = _re_fin.sub(r'[（(][^（()]{1,10}[）)]$', '', n)
    return n.strip()


def _apply_template(raw_items: list, template: list) -> list:
    """将AI提取的科目映射到标准模板，未匹配的科目追加到末尾"""
    # 构建多键查找表：原始名 / 规范名 / 别名目标 → item
    lookup: dict = {}
    for it in raw_items:
        nm = it.get("name", "").strip()
        if not nm:
            continue
        lookup[nm] = it
        nrm = _norm(nm)
        if nrm != nm:
            lookup[nrm] = it
        # 正向别名：提取名是源，别名目标作 key
        alias_target = _NAME_ALIASES.get(nm) or _NAME_ALIASES.get(nrm)
        if alias_target:
            lookup[alias_target] = it
        # 反向别名：提取名是别名目标时，也能被模板中的源名找到
        # （此处不需要反查，因为模板 key 已是标准名）

    tmpl_norms = {_norm(r["name"]) for r in template}
    matched_raw_names: set = set()
    result = []

    for row in template:
        key = row["name"]
        match = lookup.get(key) or lookup.get(_norm(key))
        if match:
            matched_raw_names.add(match.get("name", ""))
            result.append({"name": key, "current": match.get("current"),
                           "prev": match.get("prev"), "type": row["type"]})
        else:
            result.append({"name": key, "current": None, "prev": None, "type": row["type"]})

    # 未匹配的非 header 科目追加到末尾
    for it in raw_items:
        nm = it.get("name", "").strip()
        if nm and nm not in matched_raw_names and _norm(nm) not in tmpl_norms:
            if it.get("type") != "header":
                result.append({"name": nm, "current": it.get("current"),
                               "prev": it.get("prev"), "type": it.get("type", "detail")})
    return result


_TMPL_MAP = {
    "balance_sheet":          BS_CONSOL_TMPL,
    "balance_sheet_parent":   BS_PARENT_TMPL,
    "income":                 INC_CONSOL_TMPL,
    "income_parent":          INC_PARENT_TMPL,
    "cash_flow_consolidated": CF_TMPL,
    "cash_flow_parent":       CF_TMPL,
}


_STMT_KEYS = {
    "balance_sheet_consolidated": "balance_sheet",
    "balance_sheet_parent":       "balance_sheet_parent",
    "income_consolidated":        "income",
    "income_parent":              "income_parent",
    "cash_flow_consolidated":     "cash_flow_consolidated",
    "cash_flow_parent":           "cash_flow_parent",
}
_FIN_KEYWORDS = [
    "资产负债表", "利润表", "现金流量表",
    "货币资金", "营业收入", "资产总计",
    "经营活动产生", "投资活动产生", "筹资活动产生",
    "母公司资产负债", "母公司利润", "母公司现金",
]


def _read_xlsx_sheets(file_path: Path) -> dict:
    """Return {sheet_name: [[cell_value, ...]]} for all sheets in an xlsx/xlsm file."""
    import openpyxl
    wb = openpyxl.load_workbook(str(file_path), data_only=True)
    sheets = {}
    for ws in wb.worksheets:
        # Unmerge: fill merged regions with top-left value
        merge_map = {}
        for rng in list(ws.merged_cells.ranges):
            top_val = ws.cell(rng.min_row, rng.min_col).value
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    if (r, c) != (rng.min_row, rng.min_col):
                        merge_map[(r, c)] = top_val
        rows = []
        for r_idx, row in enumerate(ws.iter_rows(), start=1):
            cells = []
            for c_idx, cell in enumerate(row, start=1):
                val = merge_map.get((r_idx, c_idx), cell.value)
                cells.append(val)
            rows.append(cells)
        sheets[ws.title] = rows
    return sheets


def _read_xls_sheets(file_path: Path) -> dict:
    """Return {sheet_name: [[cell_value, ...]]} for all sheets in an xls file."""
    import xlrd
    wb = xlrd.open_workbook(str(file_path))
    sheets = {}
    for ws in wb.sheets():
        rows = []
        for r in range(ws.nrows):
            cells = []
            for c in range(ws.ncols):
                cell = ws.cell(r, c)
                if cell.ctype == xlrd.XL_CELL_EMPTY:
                    cells.append(None)
                elif cell.ctype == xlrd.XL_CELL_NUMBER:
                    v = cell.value
                    cells.append(int(v) if v == int(v) else v)
                else:
                    cells.append(cell.value)
            rows.append(cells)
        sheets[ws.name] = rows
    return sheets


def _parse_fin_sheet(rows: list) -> tuple:
    """
    Parse a 2-D table of cells into a list of financial items.
    Returns (items, unit, year_label, prev_year_label).
    items = [{"name": str, "current": float|None, "prev": float|None, "type": str}]
    """
    import re

    unit = "元"
    year_cur = ""
    year_prev = ""

    # Detect unit in first 10 rows
    for row in rows[:10]:
        for cell in row:
            s = str(cell or "")
            if "万元" in s:
                unit = "万元"
            elif "千元" in s or "千" in s:
                unit = "千元"

    # Detect name column and year columns by scanning header rows
    name_col = 0
    cur_col = None
    prev_col = None

    for row in rows[:15]:
        str_cells = [str(c or "").strip() for c in row]
        # Look for a row that contains year patterns like 2023, 2022
        years = [(i, s) for i, s in enumerate(str_cells)
                 if re.search(r'20\d{2}', s) and i > 0]
        if len(years) >= 2:
            cur_col = years[0][0]
            prev_col = years[1][0]
            year_cur = re.search(r'20\d{2}', years[0][1]).group()
            year_prev = re.search(r'20\d{2}', years[1][1]).group()
            break
        if len(years) == 1:
            cur_col = years[0][0]
            year_cur = re.search(r'20\d{2}', years[0][1]).group()

    if cur_col is None:
        # Fallback: take last two non-empty numeric columns
        for row in rows[:20]:
            num_cols = [i for i, c in enumerate(row) if isinstance(c, (int, float)) and i > 0]
            if len(num_cols) >= 2:
                cur_col = num_cols[0]
                prev_col = num_cols[1]
                break
        if cur_col is None:
            return [], unit, year_cur, year_prev

    def _to_float(v):
        if v is None:
            return None
        try:
            f = float(str(v).replace(",", "").replace("，", "").strip())
            return None if f == 0 else f
        except Exception:
            return None

    items = []
    for row in rows:
        if not row:
            continue
        name_raw = str(row[name_col] or "").strip()
        if not name_raw or len(name_raw) < 2:
            continue
        # Skip obvious header/unit rows
        if re.search(r'单位|编制|报告期|项\s*目|科\s*目|说明|附注|年度', name_raw):
            continue
        cur_val = _to_float(row[cur_col]) if cur_col < len(row) else None
        prev_val = _to_float(row[prev_col]) if prev_col is not None and prev_col < len(row) else None
        # Classify row type
        row_type = "detail"
        if re.search(r'合计|总计|净额|净利润|净增加', name_raw):
            row_type = "total"
        elif re.search(r'小计', name_raw):
            row_type = "subtotal"
        items.append({"name": name_raw, "current": cur_val, "prev": prev_val, "type": row_type})

    return items, unit, year_cur, year_prev


# Keywords used to identify which of the 6 statement types a sheet represents
_SHEET_TYPE_RULES = [
    ("cash_flow_parent",       ["母公司现金流量", "本公司现金流量"]),
    ("balance_sheet_parent",   ["母公司资产负债", "本公司资产负债", "母公司资产", "本公司资产"]),
    ("income_parent",          ["母公司利润", "本公司利润", "母公司损益", "本公司损益"]),
    ("cash_flow_consolidated", ["合并现金流量", "现金流量表", "现金流量", "合并现金", "现金及现金等价物"]),
    ("balance_sheet",          ["合并资产负债", "合并负债", "资产负债表", "资产负债"]),
    ("income",                 ["合并利润", "合并损益", "利润表", "损益表", "利润", "损益"]),
]
_SHEET_GENERIC_RULES = [
    ("cash_flow_consolidated", ["现金流", "Cash"]),
    ("balance_sheet",          ["资产", "Balance"]),
    ("income",                 ["收入", "Income", "P&L"]),
]


def _classify_sheet(sheet_name: str, rows: list = None) -> Optional[str]:
    name = sheet_name.strip()
    for key, kws in _SHEET_TYPE_RULES:
        if any(kw in name for kw in kws):
            return key
    for key, kws in _SHEET_GENERIC_RULES:
        if any(kw in name for kw in kws):
            return key
    # 名称匹配不到时，扫描前30行单元格内容判断类型
    if rows:
        text = " ".join(str(c or "") for row in rows[:30] for c in row)
        if "经营活动" in text and "现金流" in text:
            return "cash_flow_consolidated"
        if "货币资金" in text or "资产负债" in text or ("资产" in text and "负债" in text and "权益" in text):
            return "balance_sheet"
        if "营业收入" in text or "净利润" in text or "利润总额" in text:
            return "income"
    return None


_PARENT_KEYWORDS = ["母公司", "本部", "本公司", "单体", "parent"]
_CONSOLIDATED_KEYWORDS = ["合并", "consolidated", "集团"]
_PARENT_KEY_MAP = {
    "balance_sheet":          "balance_sheet_parent",
    "income":                 "income_parent",
    "cash_flow_consolidated": "cash_flow_parent",
}

def _filename_is_parent(name: str) -> Optional[bool]:
    """从文件名/sheet名判断是否母公司报表。返回 True/False/None(无法判断)。"""
    n = name.lower()
    if any(kw in n for kw in _PARENT_KEYWORDS):
        return True
    if any(kw in n for kw in _CONSOLIDATED_KEYWORDS):
        return False
    return None


def _extract_excel(file_path: Path, filename: str = "") -> dict:
    """Parse financial Excel file and return same dict structure as _extract_pages_vision."""
    suffix = file_path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        sheets = _read_xlsx_sheets(file_path)
    else:
        sheets = _read_xls_sheets(file_path)

    # 从文件名判断整体是否母公司报表
    file_is_parent = _filename_is_parent(filename or file_path.stem)

    result = {k: [] for k in _TMPL_MAP}
    result["unit"] = "万元"
    result["year"] = None
    result["prev_year"] = None

    print(f"Excel '{filename}' sheets: {list(sheets.keys())}, file_is_parent={file_is_parent}")
    assigned = set()
    for sheet_name, rows in sheets.items():
        key = _classify_sheet(sheet_name, rows)
        # 如果 key 是通用分类，根据文件名或 sheet 名决定是否映射到母公司版本
        if key in _PARENT_KEY_MAP:
            sheet_is_parent = _filename_is_parent(sheet_name)
            is_parent = sheet_is_parent if sheet_is_parent is not None else file_is_parent
            if is_parent:
                key = _PARENT_KEY_MAP[key]
        print(f"  sheet '{sheet_name}' → {key}")
        if key is None or key in assigned:
            continue
        items, unit, year_cur, year_prev = _parse_fin_sheet(rows)
        if not items:
            continue

        # 取第一个识别到年份的 sheet 作为年份
        if result["year"] is None and year_cur:
            try:
                result["year"] = int(year_cur)
            except Exception:
                pass
        if result["prev_year"] is None and year_prev:
            try:
                result["prev_year"] = int(year_prev)
            except Exception:
                pass

        def _to_wan(v, u):
            if v is None:
                return None
            if u == "元":
                return round(v / 10000, 4)
            if u == "千元":
                return round(v / 10, 4)
            return v

        converted = [
            {
                "name":    it["name"],
                "current": _to_wan(it["current"], unit),
                "prev":    _to_wan(it["prev"],    unit),
                "type":    it["type"],
            }
            for it in items if it.get("name")
        ]
        result[key].extend(converted)
        assigned.add(key)
        print(f"Excel sheet '{sheet_name}' → {key}: {len(converted)} rows, year={year_cur}")

    # Map raw items to standard templates
    for key, tmpl in _TMPL_MAP.items():
        result[key] = _apply_template(result[key], tmpl)
    return result


def _extract_text_from_pdf(path: Path) -> str:
    """从可复制文字的 PDF 直接提取文本。"""
    import fitz
    doc = fitz.open(str(path))
    parts = []
    for i, page in enumerate(doc):
        t = page.get_text().strip()
        if t:
            parts.append(f"--- 第{i+1}页 ---\n{t}")
    doc.close()
    return "\n\n".join(parts)


def _extract_text_from_excel(path: Path) -> str:
    """把 Excel 各 sheet 的单元格内容转成可检索文本。"""
    suffix = path.suffix.lower()
    sheets = _read_xlsx_sheets(path) if suffix in (".xlsx", ".xlsm") else _read_xls_sheets(path)
    parts = []
    for sheet_name, rows in sheets.items():
        parts.append(f"[{sheet_name}]")
        for row in rows:
            cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if cells:
                parts.append("  ".join(cells))
    return "\n".join(parts)


async def _extract_text_hybrid_pdf(path: Path, max_pages: int = 60) -> str:
    """逐页提取：文字页用 PyMuPDF，扫描页用 Kimi OCR，合并返回。"""
    import fitz, base64 as _b64
    api_key = os.getenv("DEEPSEEK_API_KEY", "")
    base_url = os.getenv("DEEPSEEK_BASE_URL", "")
    has_ocr = bool(api_key and "moonshot" in base_url)

    if has_ocr:
        from openai import OpenAI
        ocr_client = OpenAI(api_key=api_key, base_url=base_url)

    pdf = fitz.open(str(path))
    page_texts = []
    ocr_needed = []  # [(page_index, pixmap_b64)]

    for i in range(min(len(pdf), max_pages)):
        t = pdf[i].get_text().strip()
        if len(t) >= 50:
            page_texts.append((i, t))
        else:
            if has_ocr:
                pix = pdf[i].get_pixmap(matrix=fitz.Matrix(2.0, 2.0))
                b64 = _b64.b64encode(pix.tobytes("jpeg", jpg_quality=70)).decode()
                ocr_needed.append((i, b64))
            # 无 OCR 配置时跳过扫描页

    # OCR 扫描页（并发，最多同时 5 页，避免逐页串行累加超时）
    sem = asyncio.Semaphore(5)

    async def _ocr_page(idx: int, img_b64: str):
        async with sem:
            try:
                resp = await asyncio.to_thread(
                    ocr_client.chat.completions.create,
                    model="moonshot-v1-8k-vision-preview",
                    messages=[{"role": "user", "content": [
                        {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{img_b64}"}},
                        {"type": "text", "text": "请将这页文档的全部文字内容原文输出，保持段落结构，不添加任何解释。"},
                    ]}],
                    max_tokens=2000,
                )
                t = (resp.choices[0].message.content or "").strip()
                return (idx, t) if t else None
            except Exception as e:
                print(f"OCR page {idx+1} error: {e}")
                return None

    if ocr_needed:
        results = await asyncio.gather(*[_ocr_page(i, b64) for i, b64 in ocr_needed])
        page_texts.extend(r for r in results if r)

    pdf.close()
    page_texts.sort(key=lambda x: x[0])
    return "\n\n".join(f"--- 第{i+1}页 ---\n{t}" for i, t in page_texts)


async def _index_document(doc_id: int, file_path: Path, original_name: str):
    """上传后台任务：提取文本并写入 FTS 索引。"""
    suffix = file_path.suffix.lower()
    text = ""
    try:
        if suffix in (".xlsx", ".xls", ".xlsm"):
            text = _extract_text_from_excel(file_path)
        elif suffix == ".pdf":
            text = await _extract_text_hybrid_pdf(file_path)
        elif suffix in (".txt", ".csv", ".md"):
            text = file_path.read_text(errors="ignore")
    except Exception as e:
        print(f"Index doc {doc_id} error: {e}")

    conn = get_db()
    conn.execute(
        "UPDATE documents SET doc_text=?, doc_indexed=1 WHERE id=?",
        (text[:500_000], doc_id),
    )
    conn.execute("DELETE FROM documents_fts WHERE rowid=?", (doc_id,))
    if text.strip():
        conn.execute(
            "INSERT INTO documents_fts(rowid, original_name, doc_text) VALUES (?,?,?)",
            (doc_id, original_name, text[:500_000]),
        )
    conn.commit()
    conn.close()
    print(f"Doc {doc_id} indexed: {len(text)} chars")


def _build_fin_context(company_id: int) -> str:
    """从 company_financials 提取关键财务指标，格式化成 AI 可读的上下文。"""
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM company_financials WHERE company_id=? ORDER BY year DESC LIMIT 3",
        (company_id,),
    ).fetchall()
    conn.close()
    if not rows:
        return ""

    # 从报表科目列表中按名称取值
    def _get(items: list, *names) -> Optional[float]:
        for item in items:
            if item.get("name") in names:
                v = item.get("current")
                return v if v is not None else item.get("prev")
        return None

    KEY_METRICS = [
        ("资产总计",   ["资产总计", "资产合计"]),
        ("负债合计",   ["负债合计", "负债总计"]),
        ("所有者权益", ["所有者权益合计", "股东权益合计", "归属于母公司所有者权益合计"]),
        ("营业收入",   ["营业收入", "一、营业收入"]),
        ("净利润",     ["净利润", "七、净利润", "四、净利润"]),
        ("经营现金流", ["经营活动产生的现金流量净额"]),
    ]

    lines = ["【财务数据（单位：万元）】"]
    for row in rows:
        year = row["year"]
        unit = row.get("unit") or "万元"
        bs   = json.loads(row["balance_sheet"] or "[]")
        inc  = json.loads(row["income"] or "[]")
        cf   = json.loads(row["cash_flow_consolidated"] or "[]")
        lines.append(f"\n{year}年（合并报表）：")
        for label, names in KEY_METRICS:
            src = cf if "现金流" in label else (inc if "营业" in label or "净利润" in label else bs)
            v = _get(src, *names)
            lines.append(f"  {label}：{v:.2f} 万元" if v is not None else f"  {label}：—")
    return "\n".join(lines)


def _search_doc_chunks(company_id: int, query: str, limit: int = 4) -> list:
    """在该客户的文档 FTS 中检索与问题相关的段落。"""
    import re
    # 提取2字以上中文词组 + 4位数字（年份）
    terms = re.findall(r'[一-鿿]{2,}|[0-9]{4}', query)
    if not terms:
        return []
    # 每个词单独匹配（OR），避免短语匹配失败
    fts_query = " OR ".join(f'"{t}"' for t in terms[:12])
    conn = get_db()
    try:
        rows = conn.execute(
            """SELECT d.original_name,
                      snippet(documents_fts, 1, '【', '】', '...', 40) AS snip
               FROM documents_fts
               JOIN documents d ON d.id = documents_fts.rowid
               WHERE d.company_id = ? AND documents_fts MATCH ?
               ORDER BY rank
               LIMIT ?""",
            (company_id, fts_query, limit),
        ).fetchall()
        return [f"[{r[0]}] {r[1]}" for r in rows]
    except Exception as e:
        print(f"FTS search error: {e}")
        return []
    finally:
        conn.close()


async def _extract_pages_vision(pdf_path: Path) -> dict:
    """把PDF各页转图片，用视觉AI提取财务报表（合并+本部×3表）"""
    import fitz, base64 as _b64, re as _re
    from openai import OpenAI

    api_key = os.getenv("DEEPSEEK_API_KEY", "")
    base_url = os.getenv("DEEPSEEK_BASE_URL", "").rstrip("/")
    client = OpenAI(api_key=api_key, base_url=base_url)

    pdf = fitz.open(str(pdf_path))
    total = len(pdf)

    # 扫文字层找候选页（最多扫60页），无命中则兜底用5-35页
    candidate_pages = []
    for i in range(min(60, total)):
        txt = pdf[i].get_text() or ""
        if any(k in txt for k in _FIN_KEYWORDS):
            candidate_pages.append(i)
    if not candidate_pages:
        candidate_pages = list(range(5, min(36, total)))

    result = {"year": None, "prev_year": None, "unit": "元",
              "balance_sheet": [], "income": [],
              "balance_sheet_parent": [], "income_parent": [],
              "cash_flow_consolidated": [], "cash_flow_parent": []}

    for idx in candidate_pages:
        page = pdf[idx]
        mat = fitz.Matrix(2.0, 2.0)
        pix = page.get_pixmap(matrix=mat)
        b64 = _b64.b64encode(pix.tobytes("png")).decode()

        try:
            resp = client.chat.completions.create(
                model=VISION_MODEL,
                messages=[{"role": "user", "content": [
                    {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64}"}},
                    {"type": "text", "text": _FIN_PROMPT},
                ]}],
                max_tokens=6000,
            )
            raw = resp.choices[0].message.content
            finish = resp.choices[0].finish_reason
            m = _re.search(r'\{[\s\S]*', raw)
            if not m:
                print(f"Page {idx+1}: no JSON in response, skip")
                continue
            fragment = m.group()
            # 截断时修补残缺 JSON：找最后一个完整 item 的 }，补上 ]}
            if finish == "length":
                last_close = fragment.rfind('},')
                if last_close == -1:
                    last_close = fragment.rfind('}')
                if last_close != -1:
                    fragment = fragment[:last_close + 1] + "\n  ]\n}"
                    print(f"Page {idx+1}: truncated, salvaged up to char {last_close}")
                else:
                    print(f"Page {idx+1}: truncated and unsalvageable, skip")
                    continue
            # 确保末尾有完整 }
            if not fragment.rstrip().endswith('}'):
                last_close = fragment.rfind('}')
                fragment = fragment[:last_close + 1] if last_close != -1 else fragment
            page_data = json.loads(fragment)
        except json.JSONDecodeError as e:
            print(f"Page {idx+1}: JSON parse error: {e} — raw[:200]: {raw[:200]}")
            continue
        except Exception as e:
            print(f"Vision page {idx+1} error: {e}")
            continue

        ptype = page_data.get("type", "other")
        if ptype == "other":
            continue

        # 记录年份和单位（取第一个有效值）
        if not result["year"] and page_data.get("year"):
            result["year"] = page_data["year"]
            result["prev_year"] = page_data.get("prev_year")
        if page_data.get("unit"):
            result["unit"] = page_data["unit"]

        items = page_data.get("items", [])
        converted = [
            {
                "name": it.get("name", ""),
                "current": _to_wan(it.get("current"), page_data.get("unit", "元")),
                "prev":    _to_wan(it.get("prev"),    page_data.get("unit", "元")),
                "type":    it.get("type", "detail"),
            }
            for it in items if it.get("name")
        ]

        dest_key = _STMT_KEYS.get(ptype)
        if dest_key:
            result[dest_key].extend(converted)
            print(f"Page {idx+1}: {ptype} → {len(converted)} rows")

    pdf.close()
    result["unit"] = "万元"
    # 将原始科目列表映射到标准模板
    for key, tmpl in _TMPL_MAP.items():
        result[key] = _apply_template(result[key], tmpl)
    return result


@app.post("/api/companies/{company_id}/financials/extract")
async def extract_financial(company_id: int):
    conn = get_db()
    doc = conn.execute(
        """SELECT * FROM documents WHERE company_id=?
           AND (
             lower(original_name) LIKE '%.pdf'
             OR lower(original_name) LIKE '%.xlsx'
             OR lower(original_name) LIKE '%.xls'
             OR lower(original_name) LIKE '%.xlsm'
           )
           ORDER BY
             CASE WHEN category='财务状况' THEN 0
                  WHEN original_name LIKE '%审计%' THEN 1
                  WHEN original_name LIKE '%财务%' THEN 2
                  ELSE 3 END,
             uploaded_at DESC
           LIMIT 1""",
        (company_id,),
    ).fetchone()
    conn.close()

    if not doc:
        raise HTTPException(400, "未找到可提取的文档，请先上传审计报告（PDF）或财务报表（Excel）")

    file_path = UPLOAD_DIR / doc["filename"]
    if not file_path.exists():
        raise HTTPException(400, "文档文件不存在，请重新上传")

    suffix = Path(doc["original_name"]).suffix.lower()
    is_excel = suffix in (".xlsx", ".xls", ".xlsm")

    try:
        if is_excel:
            result = _extract_excel(file_path)
        else:
            api_key = os.getenv("DEEPSEEK_API_KEY", "")
            base_url = os.getenv("DEEPSEEK_BASE_URL", "")
            if not api_key or "moonshot" not in base_url:
                raise HTTPException(400, "需要配置 Kimi（Moonshot）云端 API 才能提取扫描件PDF")
            result = await _extract_pages_vision(file_path)
    except HTTPException:
        raise
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(500, f"提取失败：{e}")

    if not any(result[k] for k in ("balance_sheet", "income", "balance_sheet_parent",
                                    "income_parent", "cash_flow_consolidated", "cash_flow_parent")):
        raise HTTPException(400, "未识别到财务报表，请确认文档包含资产负债表、利润表或现金流量表")

    result["source_doc"] = doc["original_name"]
    return result


@app.post("/api/companies/{company_id}/financials/extract-excel")
async def extract_financial_excel(company_id: int, file: UploadFile = File(...)):
    """直接上传 Excel 报表文件，解析后返回财务数据（不保存到文档库）。"""
    ext = Path(file.filename).suffix.lower()
    if ext not in (".xlsx", ".xls", ".xlsm"):
        raise HTTPException(400, "请上传 Excel 文件（.xlsx / .xls / .xlsm）")
    tmp_path = UPLOAD_DIR / f"_tmp_{uuid.uuid4().hex}{ext}"
    try:
        tmp_path.write_bytes(await file.read())
        result = _extract_excel(tmp_path, filename=file.filename)
    except Exception as e:
        raise HTTPException(500, f"解析失败：{e}")
    finally:
        if tmp_path.exists():
            tmp_path.unlink()
    if not any(result[k] for k in ("balance_sheet", "income", "balance_sheet_parent",
                                    "income_parent", "cash_flow_consolidated", "cash_flow_parent")):
        raise HTTPException(400, "未识别到财务报表，请确认 Sheet 名称包含「资产负债」「利润」「现金流量」等关键词")
    # 表头没识别到年份时，从文件名提取
    if not result.get("year"):
        import re as _re
        m = _re.search(r'20\d{2}', file.filename)
        if m:
            result["year"] = int(m.group())
    result["source_doc"] = file.filename
    return result


# ── iCal Calendar Feed ───────────────────────────────────────────────────────

@app.get("/api/calendar/todos.ics")
def calendar_todos(token: str = ""):
    if ACCESS_PASSWORD and token != ACCESS_PASSWORD:
        from fastapi.responses import PlainTextResponse
        return PlainTextResponse("Unauthorized", status_code=401)

    from datetime import date, timedelta
    conn = get_db()
    rows = conn.execute(
        """SELECT t.*, co.name as company_name
           FROM todos t LEFT JOIN companies co ON t.company_id=co.id
           WHERE t.done=0 ORDER BY t.date"""
    ).fetchall()
    conn.close()

    def ics_date(d: str) -> str:
        return d.replace("-", "")

    def ics_date_next(d: str) -> str:
        return (date.fromisoformat(d) + timedelta(days=1)).strftime("%Y%m%d")

    priority_map = {"high": "1", "medium": "5", "low": "9"}
    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//客户档案系统//CN",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        "X-WR-CALNAME:客户待办事项",
        "X-WR-CALDESC:银行客户经理待办",
        "X-WR-TIMEZONE:Asia/Shanghai",
    ]

    for r in rows:
        t = dict(r)
        uid = f"todo-{t['id']}@client-manager"
        start = ics_date(t["date"])
        end = ics_date_next(t.get("end_date") or t["date"])
        summary = t["content"]
        if t.get("company_name"):
            summary = f"[{t['company_name']}] {summary}"
        desc_parts = []
        if t.get("company_name"):
            desc_parts.append(f"客户：{t['company_name']}")
        prio_label = {"high": "高优先级", "medium": "中优先级", "low": "低优先级"}.get(t.get("priority","medium"), "")
        if prio_label:
            desc_parts.append(prio_label)
        if t.get("end_date"):
            desc_parts.append(f"截止：{t['end_date']}")
        desc = "\\n".join(desc_parts)
        prio = priority_map.get(t.get("priority", "medium"), "5")

        lines += [
            "BEGIN:VEVENT",
            f"UID:{uid}",
            f"DTSTART;VALUE=DATE:{start}",
            f"DTEND;VALUE=DATE:{end}",
            f"SUMMARY:{summary}",
            f"DESCRIPTION:{desc}",
            f"PRIORITY:{prio}",
            "STATUS:NEEDS-ACTION",
            "END:VEVENT",
        ]

    lines.append("END:VCALENDAR")
    ics_content = "\r\n".join(lines) + "\r\n"

    from fastapi.responses import Response
    return Response(
        content=ics_content,
        media_type="text/calendar; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=todos.ics"},
    )


# ── Static ────────────────────────────────────────────────────────────────────

app.mount("/uploads", StaticFiles(directory="uploads"), name="uploads")
app.mount("/", StaticFiles(directory="static", html=True), name="static")


@app.on_event("startup")
def on_startup():
    init_db()
    init_vapid()
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
        from apscheduler.triggers.cron import CronTrigger
        scheduler = BackgroundScheduler(timezone="Asia/Shanghai")
        scheduler.add_job(send_daily_push, CronTrigger(hour=8, minute=0))
        scheduler.start()
        app.state.scheduler = scheduler
        print("✅ 每日推送已调度（08:00 Asia/Shanghai）")
    except Exception as e:
        print(f"⚠️  调度器启动失败：{e}")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=False)
